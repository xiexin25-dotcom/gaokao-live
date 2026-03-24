"""
Excel -> SQLite 迁移脚本 v2.0
将 data/ 下三个 Excel 文件导入 data/gaokao.db
完整迁移所有 66 列数据

用法：
    python -m engine.migrate_to_db          # 从项目根目录运行
    python engine/migrate_to_db.py          # 直接运行
"""
import os, sys, re, sqlite3
import openpyxl

# Windows 控制台中文输出
if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

# 路径
if getattr(sys, 'frozen', False):
    _BASE = os.path.dirname(sys.executable)
else:
    _BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

DB_PATH      = os.path.join(_BASE, 'data', 'gaokao.db')
SCHEMA_PATH  = os.path.join(_BASE, 'engine', 'schema.sql')
MAIN_XLSX    = os.path.join(_BASE, 'data', '2026_jilin_gaokao_data.xlsx')
SYBAN_XLSX   = os.path.join(_BASE, 'data', 'shiyanban.xlsx')
BENKE_XLSX   = os.path.join(_BASE, 'data', 'benkezy.xlsx')


def create_tables(conn: sqlite3.Connection):
    """执行 schema.sql 建表"""
    with open(SCHEMA_PATH, 'r', encoding='utf-8') as f:
        conn.executescript(f.read())
    print("[OK] 建表完成")


def _safe_float(v):
    """安全转 float，None/空/非数字 -> None"""
    if v is None:
        return None
    try:
        f = float(v)
        return f if f == f else None  # NaN check
    except (ValueError, TypeError):
        return None


def _safe_int(v):
    f = _safe_float(v)
    return int(f) if f is not None else None


def _safe_str(v):
    """安全转字符串，None/空/'None' -> None"""
    if v is None:
        return None
    s = str(v).strip()
    if not s or s == 'None' or s == 'nan':
        return None
    return s


def migrate_main_data(conn: sqlite3.Connection):
    """
    导入主数据 2026_jilin_gaokao_data.xlsx
    Sheet "吉林"，表头行3，数据行4+
    完整迁移所有字段
    """
    print(f"[...] 正在读取 {os.path.basename(MAIN_XLSX)} ...")
    wb = openpyxl.load_workbook(MAIN_XLSX, read_only=True, data_only=True)
    ws = wb['吉林']

    # 读取表头（第3行），处理重名
    raw_headers = [cell.value for cell in ws[3]]
    seen = {}
    headers = []
    for i, h in enumerate(raw_headers):
        if h is None:
            h = f'col_{i}'
        h = str(h).strip()
        if h in seen:
            seen[h] += 1
            headers.append(f'{h}_{seen[h]}')
        else:
            seen[h] = 0
            headers.append(h)

    # 建立列名->索引映射
    col = {h: i for i, h in enumerate(headers)}
    print(f"  表头列数: {len(headers)}")

    # 缓存：province -> id, school_key -> id, gcode -> id
    prov_cache = {}
    school_cache = {}
    group_cache = {}

    cur = conn.cursor()
    row_count = 0
    score_count = 0

    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[1] is None:
            continue  # 跳过空行

        row_count += 1

        # ── 提取所有字段 ──
        def g(name, default=None):
            """安全获取列值"""
            idx = col.get(name)
            if idx is None or idx >= len(row):
                return default
            return row[idx]

        year_val    = g('年份')
        ke_lei      = _safe_str(g('科类')) or ''
        batch       = _safe_str(g('批次')) or ''
        pub_priv    = _safe_str(g('公私性质')) or '公办'
        school_name = (_safe_str(g('院校名称')) or '').strip()
        major_name  = (_safe_str(g('专业名称')) or '').strip()
        major_full  = _safe_str(g('专业全称'))
        remark      = _safe_str(g('专业备注'))
        gcode       = (_safe_str(g('院校专业组代码')) or '').strip()
        province    = (_safe_str(g('所在省')) or '').strip()
        city        = (_safe_str(g('城市')) or '').strip()
        sch_type    = _safe_str(g('类型')) or ''
        tags        = _safe_str(g('院校标签')) or ''
        ruanke      = _safe_str(g('软科评级')) or ''
        city_level  = _safe_str(g('城市水平标签')) or ''
        subj_req    = _safe_str(g('选科要求')) or ''
        gmin_score  = _safe_float(g('专业组最低分'))
        tuition     = _safe_float(g('学费'))

        # 院校级别字段
        nat_rank    = _safe_int(g('院校排名'))
        school_lv   = _safe_str(g('院校水平'))
        admin_unit  = _safe_str(g('隶属单位'))
        bao_yan     = _safe_str(g('保研率'))
        transfer    = _safe_str(g('转专业情况'))
        master_cnt  = _safe_int(g('全校硕士专业数'))
        doctor_cnt  = _safe_int(g('全校博士专业数'))
        ruanke_rank = _safe_int(g('软科排名'))
        charter_url = _safe_str(g('2025招生章程'))

        # 专业组级别字段
        disc_eval   = _safe_str(g('学科评估'))
        group_plan  = _safe_int(g('专业组计划人数'))
        gmin_rank   = _safe_int(g('专业组最低位次'))
        admit_cnt_g = _safe_int(g('专业组录取人数'))

        # 专业级别字段
        plan_count  = _safe_int(g('计划人数'))
        study_years = _safe_int(g('学制'))
        major_gate  = _safe_str(g('门类'))
        major_class = _safe_str(g('专业类'))
        major_level = _safe_str(g('专业水平'))
        master_pt   = _safe_str(g('本专业硕士点'))
        doctor_pt   = _safe_str(g('本专业博士点'))
        is_new      = _safe_str(g('是否新增'))

        # 多年度分数列 (2025)
        s25 = _safe_float(g('最低分'))
        r25 = _safe_int(g('最低位次'))
        max25 = _safe_float(g('最高分'))
        maxr25 = _safe_int(g('最高位次'))
        adm25 = _safe_int(g('录取人数'))

        # 2024
        s24 = _safe_float(g('最低分_1'))
        r24 = _safe_int(g('最低分位次'))
        adm24 = _safe_int(g('录取人数_1'))

        # 2023
        s23 = _safe_float(g('最低分_2'))
        r23 = _safe_int(g('最低分位次_1'))
        adm23 = _safe_int(g('录取人数_2'))

        if not school_name or not gcode:
            continue

        # ── 1. province
        if province and province not in prov_cache:
            cur.execute("INSERT OR IGNORE INTO province(name, city_level) VALUES(?,?)",
                        (province, city_level))
            cur.execute("SELECT id FROM province WHERE name=?", (province,))
            prov_cache[province] = cur.fetchone()[0]
        prov_id = prov_cache.get(province)

        # ── 2. school (完整字段)
        sch_key = (school_name, city)
        if sch_key not in school_cache:
            cur.execute("""INSERT OR IGNORE INTO school
                           (name, province_id, city, type, tags, pub_priv, ruanke,
                            city_level, nat_rank, school_level, admin_unit, bao_yan,
                            transfer, master_cnt, doctor_cnt, ruanke_rank, charter_url)
                           VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (school_name, prov_id, city, sch_type, tags, pub_priv, ruanke,
                         city_level, nat_rank, school_lv, admin_unit, bao_yan,
                         transfer, master_cnt, doctor_cnt, ruanke_rank, charter_url))
            cur.execute("SELECT id FROM school WHERE name=? AND city=?", (school_name, city))
            school_cache[sch_key] = cur.fetchone()[0]
        sch_id = school_cache[sch_key]

        # ── 3. major_group (含专业组级别字段)
        year_int = _safe_int(year_val) or 2025
        grp_key = (gcode, year_int)
        if grp_key not in group_cache:
            cur.execute("""INSERT OR IGNORE INTO major_group
                           (gcode, school_id, year, ke_lei, batch, subj_req,
                            gmin_score, gmin_rank, disc_eval, group_plan, admit_cnt)
                           VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
                        (gcode, sch_id, year_int, ke_lei, batch, subj_req,
                         gmin_score, gmin_rank, disc_eval, group_plan, admit_cnt_g))
            cur.execute("SELECT id FROM major_group WHERE gcode=? AND year=?",
                        (gcode, year_int))
            group_cache[grp_key] = cur.fetchone()[0]
        grp_id = group_cache[grp_key]

        # ── 4. major_score (完整字段，每个年份一条记录)
        if s25 is not None:
            cur.execute("""INSERT INTO major_score
                           (major_group_id, major_name, major_full_name, year,
                            min_score, min_rank, max_score, max_rank,
                            tuition, plan_count, admit_count, study_years,
                            remark, major_gate, major_class, major_level,
                            master_point, doctor_point, is_new)
                           VALUES(?,?,?,2025,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (grp_id, major_name, major_full, s25, r25, max25, maxr25,
                         tuition, plan_count, adm25, study_years,
                         remark, major_gate, major_class, major_level,
                         master_pt, doctor_pt, is_new))
            score_count += 1
        if s24 is not None:
            cur.execute("""INSERT INTO major_score
                           (major_group_id, major_name, major_full_name, year,
                            min_score, min_rank, tuition, admit_count,
                            remark, major_gate, major_class)
                           VALUES(?,?,?,2024,?,?,?,?,?,?,?)""",
                        (grp_id, major_name, major_full, s24, r24, tuition, adm24,
                         remark, major_gate, major_class))
            score_count += 1
        if s23 is not None:
            cur.execute("""INSERT INTO major_score
                           (major_group_id, major_name, major_full_name, year,
                            min_score, min_rank, tuition, admit_count,
                            remark, major_gate, major_class)
                           VALUES(?,?,?,2023,?,?,?,?,?,?,?)""",
                        (grp_id, major_name, major_full, s23, r23, tuition, adm23,
                         remark, major_gate, major_class))
            score_count += 1

    wb.close()
    conn.commit()
    print(f"[OK] 主数据导入完成：{row_count} 行 Excel -> {score_count} 条 major_score 记录")
    print(f"    province: {len(prov_cache)}, school: {len(school_cache)}, major_group: {len(group_cache)}")


def migrate_syban(conn: sqlite3.Connection):
    """导入实验班映射 shiyanban.xlsx"""
    if not os.path.exists(SYBAN_XLSX):
        print("[!] shiyanban.xlsx 不存在，跳过")
        return

    print(f"[...] 正在读取 {os.path.basename(SYBAN_XLSX)} ...")
    wb = openpyxl.load_workbook(SYBAN_XLSX, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    headers = [cell.value for cell in ws[3]]

    cur = conn.cursor()
    count = 0
    for row_data in ws.iter_rows(min_row=4, values_only=True):
        d = dict(zip(headers, row_data))
        school   = str(d.get('院校名称') or '').strip()
        cls_name = str(d.get('实验班名称') or '').strip()
        full_nm  = str(d.get('实验班全称') or '').strip()
        if not school or not cls_name:
            continue

        majors = set()
        # 显式分流专业
        fen_liu = str(d.get('分流专业') or '').strip()
        if fen_liu and fen_liu != 'None':
            majors.add(fen_liu)
        # 从全称括号解析
        parts = re.findall(r'[（(]([^）)]+)[）)]', full_nm)
        for p in parts:
            if '、' in p or '，' in p:
                for item in re.split(r'[、，,]', p):
                    item = item.strip()
                    if 2 <= len(item) <= 20 and '(' not in item and '（' not in item:
                        if not any(x in item for x in ('详情', '收费', '培养', '学期', '担任', '任选')):
                            majors.add(item)

        for m in majors:
            cur.execute("""INSERT OR IGNORE INTO syban_mapping
                           (school_name, class_name, full_name, major_name)
                           VALUES(?,?,?,?)""",
                        (school, cls_name, full_nm, m))
            count += 1

    wb.close()
    conn.commit()
    print(f"[OK] 实验班映射导入完成：{count} 条记录")


def _import_catalog_sheet(conn: sqlite3.Connection, ws, level: str):
    """导入一个专业目录 sheet（本科或专科）"""
    # 自动检测表头行（找含"专业"字样的行）
    header_row = 1
    for r in range(1, 6):
        vals = [str(c.value or '') for c in ws[r]]
        if any('专业' in v for v in vals):
            header_row = r
            break

    headers = [str(cell.value or f'col_{i}') for i, cell in enumerate(ws[header_row])]

    cur = conn.cursor()
    count = 0
    for row_data in ws.iter_rows(min_row=header_row + 1, values_only=True):
        d = dict(zip(headers, row_data))
        # 尝试常见列名（本科用"门类"，专科用"大类"）
        gate = str(d.get('门类名称') or d.get('门类') or d.get('学科门类')
                   or d.get('大类') or '').strip()
        gate_code = str(d.get('门类代码') or d.get('学科门类代码')
                        or d.get('大类代码') or '').strip()
        cat = str(d.get('类别名称') or d.get('专业类') or d.get('类别') or '').strip()
        cat_code = str(d.get('类别代码') or d.get('专业类代码') or '').strip()
        major = str(d.get('专业名称') or d.get('专业') or '').strip()
        major_code = str(d.get('专业代码') or '').strip()

        if not major:
            continue

        cur.execute("""INSERT OR IGNORE INTO major_catalog
                       (level, gate, gate_code, category, cat_code, major_name, major_code)
                       VALUES(?,?,?,?,?,?,?)""",
                    (level, gate, gate_code, cat, cat_code, major, major_code))
        count += 1

    return count


def migrate_benke_catalog(conn: sqlite3.Connection):
    """导入专业目录 benkezy.xlsx（本科 + 专科）"""
    if not os.path.exists(BENKE_XLSX):
        print("[!] benkezy.xlsx 不存在，跳过")
        return

    print(f"[...] 正在读取 {os.path.basename(BENKE_XLSX)} ...")
    wb = openpyxl.load_workbook(BENKE_XLSX, read_only=True, data_only=True)

    # 导入本科（第一个 sheet）
    count_benke = _import_catalog_sheet(conn, wb.worksheets[0], '本科')
    conn.commit()
    print(f"[OK] 本科专业目录导入完成：{count_benke} 条记录")

    # 导入专科（第二个 sheet，如果存在）
    if len(wb.worksheets) >= 2:
        count_zhuanke = _import_catalog_sheet(conn, wb.worksheets[1], '专科')
        conn.commit()
        print(f"[OK] 专科专业目录导入完成：{count_zhuanke} 条记录")

    wb.close()


def verify(conn: sqlite3.Connection):
    """验证迁移结果"""
    cur = conn.cursor()
    tables = ['province', 'school', 'major_group', 'major_score',
              'major_catalog', 'syban_mapping']
    print("\n-- 验证 --")
    for t in tables:
        cur.execute(f"SELECT COUNT(*) FROM {t}")
        n = cur.fetchone()[0]
        print(f"  {t:20s} -> {n:>8,} 条")

    # 抽样检查：本科批 物理 2025 的组数
    cur.execute("""SELECT COUNT(DISTINCT mg.gcode)
                   FROM major_group mg
                   JOIN school s ON mg.school_id = s.id
                   WHERE mg.year = 2025 AND mg.ke_lei = '物理' AND mg.batch = '本科批'
                   AND s.pub_priv = '公办'""")
    n_groups = cur.fetchone()[0]
    print(f"\n  本科批/物理/公办/2025 专业组数: {n_groups}")

    # 新增校验：学科评估数据
    cur.execute("SELECT COUNT(*) FROM major_group WHERE disc_eval IS NOT NULL AND disc_eval != ''")
    n_eval = cur.fetchone()[0]
    print(f"  含学科评估数据的专业组数: {n_eval}")

    # 新增校验：专项计划
    cur.execute("SELECT COUNT(*) FROM major_score WHERE remark LIKE '%专项%' AND year=2025")
    n_zx = cur.fetchone()[0]
    print(f"  含专项计划标记的2025专业数: {n_zx}")

    # 新增校验：院校附加信息
    cur.execute("SELECT COUNT(*) FROM school WHERE bao_yan IS NOT NULL")
    n_by = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM school WHERE school_level IS NOT NULL")
    n_lv = cur.fetchone()[0]
    print(f"  含保研率的院校数: {n_by}, 含院校水平的: {n_lv}")

    # 新增校验：专业附加信息
    cur.execute("SELECT COUNT(*) FROM major_score WHERE major_level IS NOT NULL AND year=2025")
    n_ml = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM major_score WHERE master_point IS NOT NULL AND year=2025")
    n_mp = cur.fetchone()[0]
    print(f"  含专业水平的: {n_ml}, 含硕士点的: {n_mp}")

    # 前5所学校
    cur.execute("""SELECT s.name, s.city, s.tags, s.bao_yan, COUNT(mg.id) as n_groups
                   FROM school s
                   JOIN major_group mg ON mg.school_id = s.id
                   WHERE mg.batch = '本科批'
                   GROUP BY s.id ORDER BY n_groups DESC LIMIT 5""")
    print("\n  Top 5 院校（按专业组数）：")
    for r in cur.fetchall():
        by = r[3] or '-'
        print(f"    {r[0]:20s}  {r[1]:8s}  {r[2][:20]:20s}  保研={by:8s}  组数={r[4]}")


def main():
    # 删除旧数据库（全量重建）
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
        print(f"[...] 已删除旧数据库 {DB_PATH}")

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")

    try:
        create_tables(conn)
        migrate_main_data(conn)
        migrate_syban(conn)
        migrate_benke_catalog(conn)
        verify(conn)
        print(f"\n[OK] 数据库迁移完成 -> {DB_PATH}")
        print(f"    文件大小: {os.path.getsize(DB_PATH) / 1024 / 1024:.1f} MB")
    finally:
        conn.close()


if __name__ == '__main__':
    main()
