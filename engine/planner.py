"""
吉林省高考志愿规划引擎 v3
核心：根据考生信息从数据源生成40志愿方案（10冲+20稳+10保）
"""
import os, sys, re, pickle, random
import pandas as pd
from engine.sybandb import is_syban_target, matching_majors as syban_matching
# syban 数据加载：SQLite 优先，回退到 Excel
from engine.sybandb import load_syban_map as _syban_load_excel

# PyInstaller 打包路径兼容
if getattr(sys, 'frozen', False):
    _DATA_BASE = os.path.dirname(sys.executable)
else:
    _DATA_BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_PATH  = os.path.join(_DATA_BASE, 'data', '2026_jilin_gaokao_data.xlsx')
CACHE_PATH = os.path.join(_DATA_BASE, 'data', 'df_cache.pkl')
DB_PATH    = os.path.join(_DATA_BASE, 'data', 'gaokao.db')

# 优先使用 SQLite 数据库
_USE_DB = os.path.exists(DB_PATH)

def load_syban_map():
    """实验班映射：SQLite 优先，回退到 Excel"""
    if _USE_DB:
        from engine.db import load_syban_map as _db_syban
        result = _db_syban()
        if result:
            return result
    return _syban_load_excel()

LV_LABEL = {1:'985', 2:'211+', 3:'211', 4:'国重点', 5:'省重点', 6:'其他'}
CR_LABEL  = {1:'一线', 2:'新一线', 3:'二线', 4:'其他'}

def school_level(tag):
    tag = str(tag) if tag else ''
    if '985' in tag: return 1
    if '211' in tag and '双一流' in tag: return 2
    if '211' in tag or '双一流' in tag: return 3
    if '国重点' in tag: return 4
    if '省重点' in tag: return 5
    return 6

def city_rank(tag):
    tag = str(tag) if tag else ''
    if '一线' in tag and '新一线' not in tag: return 1
    if '新一线' in tag: return 2
    if '二线' in tag or '省会' in tag: return 3
    return 4

def ruanke_lv(r):
    return {'A+':1,'A':2,'B+':3,'B':4}.get(str(r), 5)

KW_LIB = {
    'CS_AI':   ['计算机','人工智能','软件','信息安全','网络工程','网络空间','智能科学','数据科学','大数据','物联网','微电子','集成电路'],
    'ELEC':    ['电子信息','通信','电气','自动化','电子科学','光电','机器人','雷达'],
    'MATH':    ['数学','统计学','数学类','统计学类','精算','信息与计算科学'],
    'FINANCE': ['金融','经济','财政','保险','投资','国际贸易','财务管理'],
    'MECH':    ['机械','制造','工业工程','车辆','航空','动力工程','新能源'],
    'ARCH':    ['建筑学','城乡规划','风景园林','室内设计'],
    'MED':     ['临床医学','口腔医学','药学','中西医','中药','医学影像','公共卫生'],
    'LAW':     ['法学','法律','知识产权','国际法','司法'],
    'TEACH':   ['教育学','小学教育','学前教育','特殊教育','教育技术'],
    'MGMT':    ['工商管理','行政管理','公共管理','物流管理','人力资源'],
    'LANG':    ['英语','日语','翻译','汉语言','新闻','传播','广告','播音'],
    'ART':     ['美术','设计','音乐','舞蹈','表演','影视','数字媒体艺术'],
    'AGRI':    ['农学','植物','园艺','动物','林学','水产','食品科学'],
    'CHEM':    ['化工','化学','应用化学','高分子','材料科学','冶金'],
    'GEO':     ['地质','地理','测绘','采矿','土地资源','海洋'],
    'CIVIL':   ['土木','交通运输','道路桥梁','水利','港口航道'],
}

EXCLUDE_PRESETS = {
    'nursing':  ['护理','助产'],
    'biochem':  ['生物工程','生化','化工','应用化学','高分子','发酵工程','生物制药'],
    'medicine': ['临床医学','口腔医学','药学','中西医','中药','医学影像','麻醉','基础医学','护理'],
    'agri':     ['农学','植物科学','园艺','动物科学','林学','水产','食品科学','农业'],
    'civil':    ['土木工程','建筑工程','测绘','采矿','地质','水利','港口','船舶','道路桥梁','给排水'],
    'law':      ['法学','法律','诉讼','司法','法医'],
    'art':      ['美术','设计艺术','音乐','舞蹈','表演'],
    'finance':  ['金融','经济学','财政','保险','投资学','国际贸易','会计'],
    'chem_mat': ['化工','化学工程','应用化学','高分子','材料科学','冶金','采矿'],
    'lang':     ['英语','日语','德语','法语','翻译','外国语'],
}

_df_cache_fallback = None  # 仅用于非 DB 回退路径

def load_raw_df():
    global _df_cache_fallback
    # ── 优先从 SQLite 读取（速度快、结构化，db.py 内部有缓存） ──
    if _USE_DB:
        from engine.db import load_raw_df as _db_load
        return _db_load()

    if _df_cache_fallback is not None:
        return _df_cache_fallback

    # ── 回退：pickle 缓存 ──
    if os.path.exists(CACHE_PATH):
        with open(CACHE_PATH, 'rb') as f:
            _df_cache_fallback = pickle.load(f)
        return _df_cache_fallback

    # ── 回退：直读 Excel（首次运行 / 无DB时） ──
    import openpyxl
    wb = openpyxl.load_workbook(DATA_PATH, read_only=True, data_only=True)
    ws = wb['吉林']
    headers = [cell.value for cell in ws[3]]
    seen = {}; new_headers = []
    for i, h in enumerate(headers):
        if h is None: h = f'col_{i}'
        if h in seen: seen[h] += 1; new_headers.append(f'{h}_{seen[h]}')
        else: seen[h] = 0; new_headers.append(h)
    data = [row for row in ws.iter_rows(min_row=4, values_only=True) if row[1] is not None]
    df = pd.DataFrame(data, columns=new_headers)
    with open(CACHE_PATH, 'wb') as f:
        pickle.dump(df, f)
    _df_cache_fallback = df
    wb.close()
    return df

def build_plan(profile: dict) -> dict:
    score      = int(profile['score'])
    ke_lei     = profile.get('ke_lei', '物理')
    if ke_lei not in ('物理', '历史'):
        raise ValueError(f"科类必须为'物理'或'历史'，收到: {ke_lei!r}")
    target_kw  = profile.get('target_kw', [])
    exclude_kw = profile.get('exclude_kw', [])
    strict_exclude  = profile.get('strict_exclude', False)   # 严格排除：含排除专业的组整体移除
    exclude_ne      = profile.get('exclude_northeast', False)
    pref_provinces  = profile.get('pref_provinces', [])    # 白名单（包含）
    exc_provinces   = profile.get('exclude_provinces', []) # 黑名单（排除）
    inc_types       = profile.get('include_types', [])     # 院校类型白名单
    exc_types       = profile.get('exclude_types', [])     # 院校类型黑名单
    fee_max         = profile.get('fee_max', None)          # 学费上限（元），None=不限
    min_cr          = int(profile.get('min_city_rank', 4))
    school_pref     = profile.get('school_pref', 'school')
    slope           = float(profile.get('slope', 150.0))
    # P6：细化选科要求（3+1+2新高考，物理方向可进一步指定化学/生物/地理等）
    # 格式：['化学', '生物'] 表示要求"必须含化学且必须含生物"的专业组才保留
    # 空列表=不限（默认），包含'不限'=仅保留选科要求为"不限"的专业组
    select_subjects  = profile.get('select_subjects', [])   # 用户已选的再选科
    student_province = profile.get('student_province', '吉林')
    batch = profile.get('batch') or '本科批'          # 批次（默认本科批）
    student_rank = max(1, int(7806 + (585 - score) * slope))

    df = load_raw_df()
    if df.empty or '最低分' not in df.columns:
        return {'plan_vols': [], 'stats': {}, 'profile': profile,
                'mode': 'group', 'all_results': {}}
    _syd = df['生源地'] if '生源地' in df.columns else None
    d  = df[(df['年份'] == 2025) &
            (df['科类'] == ke_lei) &
            (df['批次'] == batch) &
            (df['公私性质'] == '公办') &
            (_syd == student_province if _syd is not None else True)].copy()
    if d.empty:
        return {'plan_vols': [], 'stats': {}, 'profile': profile,
                'mode': 'group', 'all_results': {}}

    # ── P6 选科要求细化过滤 ─────────────────────────────────────────────
    # 数据列"选科要求"示例值："物理,化学" / "物理,生物" / "不限" / None
    # 过滤规则：若用户指定了 select_subjects，则只保留：
    #   ① 选科要求为空/NaN/不限（对任意选科开放）
    #   ② 选科要求中包含用户所有指定科目（交集校验）
    if select_subjects and '选科要求' in d.columns:
        def _subj_ok(req_val):
            if not req_val or (isinstance(req_val, float) and pd.isna(req_val)):
                return True   # 无要求，开放
            req_str = str(req_val).strip()
            if req_str in ('不限', '无', ''):
                return True   # 明确标注不限
            # 将要求拆分为科目集合（支持逗号、顿号、空格分隔）
            req_set = set(re.split(r'[，,、\s]+', req_str))
            req_set.discard('')
            # 用户选科必须覆盖全部要求科目
            user_set = set(select_subjects)
            # 物理/历史已由 ke_lei 确定，无需再检查
            # 只校验再选科（化学/生物/地理/政治/历史/物理 中的额外要求）
            extra_req = req_set - {'物理', '历史'}
            return extra_req.issubset(user_set)
        mask = d['选科要求'].apply(_subj_ok)
        # 按专业组聚合：只要组内有一行不满足选科，整个专业组排除
        # 实际操作：先按专业行过滤，过滤后组内可选专业减少
        d = d[mask]
    # ────────────────────────────────────────────────────────────────────

    d['s25']    = pd.to_numeric(d['最低分'],       errors='coerce')
    d['s24']    = pd.to_numeric(d['最低分_1'],     errors='coerce')
    d['s23']    = pd.to_numeric(d['最低分_2'],     errors='coerce')
    d['r25']    = pd.to_numeric(d['最低位次'],      errors='coerce')   # 2025 专业最低位次
    d['r24']    = pd.to_numeric(d['最低分位次'],    errors='coerce')   # 2024 专业最低位次
    d['gmin25'] = pd.to_numeric(d['专业组最低分'],  errors='coerce')
    d['fee']    = pd.to_numeric(d['学费'],          errors='coerce').fillna(0)
    d['school_lv'] = d['院校标签'].apply(school_level)
    d['ruanke_lv'] = d['软科评级'].apply(ruanke_lv)
    d['city_rank']  = d['城市水平标签'].apply(city_rank)

    # 内置就业差/冷门专业关键词（独立于用户 exclude_kw，始终生效）
    BUILTIN_COLD = ['合成生物','环境科学','环境工程','大气科学','地球物理',
                    '天文','水产','草业','考古','古生物','地质']

    # 关键词同时匹配专业名称和院校名称
    # 例如"师范"匹配不到专业名，但能匹配"北京师范大学"等院校名
    # 第三路：实验班覆盖目标专业（如"工科试验班"涵盖"电气工程"）
    _syban_map = load_syban_map()

    def classify(row):
        major  = str(row['专业名称'])
        school = str(row['院校名称'])
        # 1) 用户显式排除 —— 最高优先级
        if any(k in major  for k in exclude_kw):   return 'cold'
        # 2) 用户显式目标 —— 优先于内置冷门（用户指定"环境科学"不应被覆盖）
        if any(k in major  for k in target_kw):    return 'target'
        if any(k in school for k in target_kw):    return 'target'  # 院校名兜底
        # 实验班：检查该(院校,专业名)是否是覆盖目标专业的实验班
        if target_kw and _syban_map and (school, major) in _syban_map:
            if is_syban_target(school, major, target_kw):
                return 'target'
        # 3) 内置冷门 —— 仅对用户未指定的专业生效
        if any(k in major  for k in BUILTIN_COLD): return 'cold'
        return 'other'
    d['kind'] = d.apply(classify, axis=1)

    if pref_provinces:
        d = d[d['所在省'].isin(pref_provinces)]
    else:
        # 黑名单排除（可与 exclude_ne 叠加）
        excl = list(exc_provinces)
        if exclude_ne:
            excl += ['吉林','辽宁','黑龙江']
        if excl:
            d = d[~d['所在省'].isin(excl)]

    # ── 院校类型过滤（匹配 类型 字段，支持多值如"综合 师范"）──
    if inc_types:
        mask = d['类型'].apply(lambda t: any(tp in str(t) for tp in inc_types))
        d = d[mask]
    if exc_types:
        mask = d['类型'].apply(lambda t: any(tp in str(t) for tp in exc_types))
        d = d[~mask]

    if min_cr < 4:
        d = d[d['city_rank'] <= min_cr]

    # 学费过滤：专业组内所有专业的学费均 ≤ fee_max（NaN视为0，不过滤）
    if fee_max is not None:
        # 先算出每个专业组代码的最高学费
        grp_fee_max = (d.groupby('院校专业组代码')['fee']
                        .max().fillna(0).rename('grp_fee_max'))
        d = d.join(grp_fee_max, on='院校专业组代码')
        d = d[(d['grp_fee_max'] == 0) | (d['grp_fee_max'] <= fee_max)]
        d = d.drop(columns=['grp_fee_max'], errors='ignore')

    # 逐行聚合专业组
    groups = {}
    for _, row in d.iterrows():
        gc = row['院校专业组代码']
        if gc not in groups:
            groups[gc] = {
                'gcode': gc, 'school': row['院校名称'],
                'city': row['城市'], 'city_lv': row['城市水平标签'],
                'gmin25': row['gmin25'], 'school_lv': row['school_lv'],
                'ruanke_lv': row['ruanke_lv'], 'city_rank': row['city_rank'],
                'tag': row['院校标签'], 'province': row['所在省'],
                'majors': [], 'n_target': 0, 'n_cold': 0,
                'gmin24': None, 'gmin23': None,   # 由专业级推算
            }
        g = groups[gc]
        _major_name = str(row['专业名称'])
        _school     = str(row['院校名称'])
        # 若该条目是实验班：记录命中目标的分流专业（高亮用）+ 全量分流专业（展示用）
        _is_syban = _syban_map and (_school, _major_name) in _syban_map
        _syban_hits = syban_matching(_school, _major_name, target_kw) if (_is_syban and target_kw) else []
        _syban_all  = sorted(_syban_map[(_school, _major_name)]) if _is_syban else []
        # 检测专项计划标记（国家专项/高校专项/地方专项）
        _remark = str(row.get('专业备注', '') or '')
        _full_name = str(row.get('专业全称', '') or '')
        _zx_text = _remark or _full_name
        _zx_type = ''
        if '国家专项计划' in _zx_text:
            _zx_type = '国家专项'
        elif '高校专项计划' in _zx_text:
            _zx_type = '高校专项'
        elif '地方专项计划' in _zx_text:
            _zx_type = '地方专项'

        g['majors'].append({
            'name': _major_name, 's25': row['s25'],
            's24': row['s24'],   's23': row['s23'],
            'fee': row['fee'],   'kind': row['kind'],
            'r25': (int(row['r25']) if pd.notna(row.get('r25', float('nan'))) else None),
            'r24': (int(row['r24']) if pd.notna(row.get('r24', float('nan'))) else None),
            'syban_majors': _syban_hits,  # 命中目标专业（高亮）
            'syban_all':    _syban_all,   # 全量分流专业
            'zhuanxiang': _zx_type,       # 专项计划类型（空=非专项）
            'full_name': _full_name if _full_name != _major_name else '',
        })
        # 推算组级多年最低分（取组内所有专业的年度最低值）
        s24v = row['s24'] if pd.notna(row['s24']) else None
        s23v = row['s23'] if pd.notna(row['s23']) else None
        if s24v is not None:
            g['gmin24'] = min(g['gmin24'], s24v) if g['gmin24'] is not None else s24v
        if s23v is not None:
            g['gmin23'] = min(g['gmin23'], s23v) if g['gmin23'] is not None else s23v
        if row['kind'] == 'target': g['n_target'] += 1
        if row['kind'] == 'cold':   g['n_cold']   += 1

    def s2r(s):
        if s is None or (isinstance(s, float) and pd.isna(s)): return 999999
        return max(1, int((score - float(s)) * slope + student_rank))

    rows = []
    for gc, g in groups.items():
        gmin = g['gmin25']
        if gmin is None or (isinstance(gmin, float) and pd.isna(gmin)): continue
        # 有 target_kw 时要求组内至少含一个目标专业；无 target_kw 时不过滤（全部专业均视为可选）
        if target_kw and g['n_target'] == 0: continue

        # 用组内 target 专业最低分作为范围基准（避免混合专业组把 gmin25 拉偏）
        # 冲稳上限：+25分；保底下限：-150分（允许大差值保底，教育/艺术类专业常见）
        # 当 target_kw=[] 时，所有非 cold 专业均视为 target
        all_valid = [m for m in g['majors'] if m['s25'] and not pd.isna(m['s25']) and m['kind'] != 'cold']
        target_s25_vals = [m['s25'] for m in g['majors']
                           if m['kind'] == 'target' and m['s25'] and not pd.isna(m['s25'])]
        ref_vals = target_s25_vals if target_s25_vals else [m['s25'] for m in all_valid]
        ref = min(ref_vals) if ref_vals else float(gmin)
        if ref < score - 150 or ref > score + 30: continue

        # 当用户指定 target_kw 时，若组内所有冷门专业的分数均低于目标专业最低分，
        # 则考生凭目标专业分数录取后不会被调剂到冷门，此时记录 n_cold_over_ref=0 允许入围。
        n_cold_over_ref = sum(
            1 for m in g['majors']
            if m.get('kind') == 'cold'
            and m.get('s25') is not None and not (isinstance(m['s25'], float) and pd.isna(m['s25']))
            and m['s25'] > ref
        )

        # intent：有 target_kw 时取 target 专业，否则取所有非 cold 专业
        if target_kw:
            intent = sorted(
                [m for m in g['majors'] if m['kind'] == 'target' and m['s25'] and not pd.isna(m['s25'])],
                key=lambda m: -m['s25']
            )
        else:
            intent = sorted(all_valid, key=lambda m: -m['s25'])
        if not intent: continue

        # 铁律一：差>30分的专业基本无法录取，不放入top6
        MAX_DIFF = 30
        intent_filtered = [m for m in intent if m['s25'] <= score + MAX_DIFF]
        if not intent_filtered:
            intent_filtered = intent[:1]   # 至少保留一个（最低分的target专业）

        # 补充 other 类专业（分数 ≤ 考生分，未出现在 intent 中）
        intent_names = {m['name'] for m in intent_filtered}
        others = sorted(
            [m for m in g['majors'] if m['kind'] == 'other'
             and m['name'] not in intent_names
             and m['s25'] and not pd.isna(m['s25']) and m['s25'] <= score],
            key=lambda m: -m['s25']
        )

        # BUG-01修复：合并后全局降序，保证铁律二（组内 s25 不升序）
        candidates = intent_filtered + others
        # BUG-04修复：按专业名去重，保留 s25 更高的那条，记录去重数量
        name_to_major: dict = {}
        for m in candidates:
            nm = m['name']
            if nm not in name_to_major:
                name_to_major[nm] = m
            elif (m.get('s25') or 0) > (name_to_major[nm].get('s25') or 0):
                name_to_major[nm] = m   # 同名取高分版本
        dedup_count = len(candidates) - len(name_to_major)
        unique_cands = list(name_to_major.values())

        # 全局按 s25 降序，相同分数时 target 专业优先
        unique_cands.sort(key=lambda m: (-m['s25'], 0 if m['kind'] == 'target' else 1))

        # 目标专业置顶逻辑：
        #   情形A：用户指定 target_kw 且目标专业已达录取线（s25 ≤ score）
        #     → 始终提至①位，移除所有 s25 > top_intent['s25'] 的非目标专业，维持铁律二降序。
        #     这是必要的：若把高分 other 专业置①，考生会被分配到该专业而非目标专业。
        #   情形B：无 target_kw 或目标专业超出考生分
        #     → 仅在与首位差 ≤ 3分 且不破坏降序时才提升（原有逻辑）
        top_intent = intent_filtered[0] if intent_filtered else None
        if top_intent and len(unique_cands) >= 1 and unique_cands[0]['name'] != top_intent['name']:
            if (target_kw
                    and top_intent['s25'] is not None
                    and top_intent['s25'] <= score):
                # 情形A：强制置①，过滤掉分数高于目标专业的非目标候选
                rest = [m for m in unique_cands
                        if m['name'] != top_intent['name']
                        and (m['s25'] is None or m['s25'] <= top_intent['s25'])]
                unique_cands = [top_intent] + rest
            elif (top_intent['s25'] is not None
                    and unique_cands[0]['s25'] is not None
                    and top_intent['s25'] >= unique_cands[0]['s25'] - 3):
                # 情形B：3分内微调，仍需保证降序
                rest = [m for m in unique_cands if m['name'] != top_intent['name']]
                second_s25 = rest[0]['s25'] if rest else 0
                if top_intent['s25'] >= (second_s25 or 0):
                    unique_cands = [top_intent] + rest

        # 补满到6个：从组内所有未选专业（含 cold）中按优先级填满
        if len(unique_cands) < 6:
            selected_names = {m['name'] for m in unique_cands}
            fill_pool = [m for m in g['majors']
                         if m['name'] not in selected_names
                         and m.get('s25') and not pd.isna(m['s25'])]
            # 非cold优先，同优先级内 s25 降序
            fill_pool.sort(key=lambda m: (1 if m.get('kind') == 'cold' else 0, -m['s25']))
            needed = 6 - len(unique_cands)
            unique_cands = unique_cands + fill_pool[:needed]
            # 重新全局排序维护铁律二
            unique_cands.sort(key=lambda m: (-m['s25'], 0 if m['kind'] == 'target' else 1))

        top6 = unique_cands[:6]

        # 断言铁律二（调试用：违反时抛出 ValueError）
        for _i in range(len(top6) - 1):
            if top6[_i]['s25'] is not None and top6[_i+1]['s25'] is not None:
                if top6[_i]['s25'] < top6[_i+1]['s25']:
                    raise ValueError(
                        f"铁律二违反：{g['school']} "
                        f"top6[{_i}]={top6[_i]['name']}({top6[_i]['s25']}) "
                        f"< top6[{_i+1}]={top6[_i+1]['name']}({top6[_i+1]['s25']})"
                    )

        sc6  = top6[-1]['s25'] if top6 else None
        safe = bool(sc6 is not None and score >= sc6)

        for m in intent:
            m['rank25'] = s2r(m['s25'])
            m['diff']   = round(m['s25'] - score, 1) if m['s25'] else None
        for m in top6:
            m['rank25'] = s2r(m['s25'])
            m['diff']   = round(m['s25'] - score, 1) if m['s25'] else None

        # 规则3：组内专业分差>15 且含非target专业 → 记录风险分
        all_s25 = [m['s25'] for m in g['majors'] if m['s25'] and not pd.isna(m['s25'])]
        major_spread = (max(all_s25) - min(all_s25)) if len(all_s25) >= 2 else 0
        # 风险惩罚：spread>15 时给 sort_key 一个惩罚（0=正常，1=有风险）
        spread_risk = 1 if (major_spread > 15 and g['n_cold'] > 0) else 0

        # 记录组内命中用户 exclude_kw 的专业名称（不含内置冷门，仅用户主动排除的）
        excl_in_group = [
            m['name'] for m in g['majors']
            if any(k in m['name'] for k in exclude_kw)
        ] if exclude_kw else []

        # 检测该组是否含专项计划专业
        zx_types_in_group = list({m['zhuanxiang'] for m in g['majors'] if m.get('zhuanxiang')})

        rows.append({
            **{k: v for k, v in g.items() if k != 'majors'},
            'majors': g['majors'], 'intent': intent, 'top6': top6,
            'sc6': sc6, 'safe': safe,
            'dedup_count': dedup_count,
            'gmin_rank': s2r(gmin),
            'lv_label': LV_LABEL.get(int(g['school_lv']), '?'),
            'cr_label': CR_LABEL.get(int(g['city_rank']), '?'),
            'major_spread':    round(major_spread, 1),
            'spread_risk':     spread_risk,
            'n_cold_over_ref': n_cold_over_ref,
            'excl_in_group':   excl_in_group,   # 用户排除专业中出现在本组的列表
            'has_zhuanxiang':  bool(zx_types_in_group),  # 是否含专项计划
            'zhuanxiang_types': zx_types_in_group,       # 专项类型列表
        })

    # ── 严格排除模式：含排除专业的组整体移除 ──
    if strict_exclude and exclude_kw:
        rows = [r for r in rows if not r.get('excl_in_group')]

    all_results = {r['gcode']: r for r in rows}

    def _gmin(r):
        """安全取 gmin25 浮点值，None/NaN → 0"""
        v = r.get('gmin25')
        if v is None: return 0.0
        try: return float(v)
        except (TypeError, ValueError): return 0.0

    if school_pref == 'city':
        def sort_key(r): return (r.get('spread_risk',0), r['city_rank'], r['school_lv'], r['ruanke_lv'], -_gmin(r))
    else:
        def sort_key(r): return (r.get('spread_risk',0), r['school_lv'], r['city_rank'], r['ruanke_lv'], -_gmin(r))

    # Bug①修复：加入提档可达性校验（gmin_rank×1.15 ≥ student_rank）
    # Bug③修复：冲志愿独立排序——先按gmin接近分数线（超分越少越前），同等超分再按院校层次
    def rush_sort_key(r):
        gmin_diff = _gmin(r) - score          # 超分越少越好（越接近越前）
        return (gmin_diff, r['school_lv'], r['city_rank'], r['ruanke_lv'])

    # 冲志愿候选：按超分区间 (score, score+60]，不做 gmin_rank 过滤
    # 【吉林平行志愿录取规则】
    # - 服从调剂=True（强制）→ 永不退档。退档仅发生在不服从调剂时，系统强制服从。
    # - sc6 ≤ 考生分的作用：防被调剂到未填报专业。
    #   若score < sc6，①-⑥全部达不到 → 调剂触发 → 学校可分配组内任意专业（含未填报的冷门）。
    #   因此要求 sc6 ≤ 考生分，确保至少录入⑥（考生主动选择的保底专业），而非未知专业。
    # 冲志愿：gmin25 > score，因此组内所有专业 s25 >= gmin25 > score，
    # sc6 必然 > score，不能加 sc6<=score 条件（否则永远0冲）。
    # 冲区的"调剂风险"通过 warn_few_majors/warn_critical 提示即可。
    def _cold_ok(r):
        """n_cold==0，或 target_kw 已指定且组内所有冷门专业分数均低于目标专业（不会被调剂到冷门）"""
        return r['n_cold'] == 0 or (target_kw and r.get('n_cold_over_ref', 0) == 0)

    rush_cands = sorted([r for r in rows
        if _gmin(r) > score
        and _gmin(r) <= score + 60
        and _cold_ok(r)
    ], key=rush_sort_key)

    # 冲区不足8个时自动扩展候选范围到 score+80
    if len(rush_cands) < 8:
        used_gcodes_rush = {r['gcode'] for r in rush_cands}
        extra_rush = sorted([r for r in rows
            if _gmin(r) > score + 60
            and _gmin(r) <= score + 80
            and _cold_ok(r)
            and r['gcode'] not in used_gcodes_rush
        ], key=rush_sort_key)
        rush_cands = rush_cands + extra_rush

    safe_cands = sorted([r for r in rows
        if r['sc6'] and score - 45 <= r['sc6'] <= score - 2   # [score-45, score-2]：稳区有效区间
        and _gmin(r) <= score and _cold_ok(r)
    ], key=sort_key)

    BAO_SC6_FLOOR = score - 50   # 保区 sc6 下限：差距超 50 分几乎不会被选中，视为冗余槽位
    bao_cands = sorted([r for r in rows
        if r['sc6'] and 10 <= score - r['sc6'] <= 50    # 差距 10~50 分：真正的保区
        and _gmin(r) > 0 and _gmin(r) <= score - 5
        and _cold_ok(r)
    ], key=sort_key)

    used_schools = set(); used_codes = set()

    def pick(cands, n):
        """基础取样，全局去重（同校只取一次）"""
        res = []
        for r in cands:
            if len(res) >= n: break
            if r['gcode'] in used_codes or r['school'] in used_schools: continue
            res.append(r)
            used_schools.add(r['school']); used_codes.add(r['gcode'])
        return res

    def pick_tiered(cands, n, tier_size=8, max_per_tier=3):
        """按 sc6 分层取样，保证梯度均匀覆盖，同时按质量优先"""
        if not cands: return []
        sc6_vals = [r.get('sc6') or 0 for r in cands if r.get('sc6')]
        if not sc6_vals: return pick(cands, n)
        sc6_max = max(sc6_vals)
        tier_counts = {}
        res = []
        for r in cands:
            if len(res) >= n: break
            if r['gcode'] in used_codes or r['school'] in used_schools: continue
            sc6 = r.get('sc6') or 0
            tier = int((sc6_max - sc6) / tier_size)
            if tier_counts.get(tier, 0) >= max_per_tier: continue
            res.append(r)
            used_codes.add(r['gcode']); used_schools.add(r['school'])
            tier_counts[tier] = tier_counts.get(tier, 0) + 1
        # 若分层后仍不足，放开tier限制补充
        if len(res) < n:
            for r in cands:
                if len(res) >= n: break
                if r['gcode'] in used_codes or r['school'] in used_schools: continue
                res.append(r)
                used_codes.add(r['gcode']); used_schools.add(r['school'])
        return res

    # ── 志愿上限：根据批次 slots 动态调整 ──
    # 默认本科批 40 (冲10+稳20+保10)，提前批/专科批按 slots 等比缩放
    _batch_info = _PROVINCE_BATCHES.get(student_province, [])
    _batch_slots = next((b['slots'] for b in _batch_info if b['key'] == batch), 40)
    if _batch_slots >= 40:
        RUSH_N  = 10
        STABLE_N = 20
        SAFE_N  = 10
    else:
        # 等比缩放：冲25% + 稳50% + 保25%
        RUSH_N   = max(1, round(_batch_slots * 0.25))
        STABLE_N = max(1, round(_batch_slots * 0.50))
        SAFE_N   = max(1, _batch_slots - RUSH_N - STABLE_N)

    # 冲志愿：按超分接近优先排序，不做梯度分层
    rush_sel = pick(rush_cands, RUSH_N)

    # 稳志愿：按sc6分层取样（每8分一档，每档最多4个），保证梯度覆盖20格
    safe_sel = pick_tiered(safe_cands, STABLE_N, tier_size=8, max_per_tier=4)
    if len(safe_sel) < 14:
        extra = sorted([r for r in rows
            if r['sc6'] and score - 45 <= r['sc6'] <= score - 2   # 稳区补充也限制下限，防止超低sc6学校溢出
            and _gmin(r) > 0 and _gmin(r) <= score
            and r['n_cold'] == 0
            and r['gcode'] not in used_codes], key=sort_key)
        safe_sel += pick(extra, STABLE_N - len(safe_sel))

    # 保志愿：同样分层取样
    bao_sel = pick_tiered(bao_cands, SAFE_N, tier_size=8, max_per_tier=3)
    if len(bao_sel) < 6:
        extra = sorted([r for r in rows
            if r['sc6'] and 5 <= score - r['sc6'] <= 55   # 与 bao_cands 放宽5分
            and _gmin(r) > 0 and _gmin(r) <= score - 5
            and r['n_cold'] == 0
            and r['gcode'] not in used_codes], key=sort_key)
        bao_sel += pick(extra, SAFE_N - len(bao_sel))

    # 规则2：稳/保区内按 sc6 高→低排列（最好的院校先被系统投档匹配）
    safe_sel_sorted = sorted(safe_sel[:STABLE_N], key=lambda r: -(r.get('sc6') or 0))
    bao_sel_sorted  = sorted(bao_sel[:SAFE_N],  key=lambda r: -(r.get('sc6') or 0))

    # 稳/保分区独立排序：稳区先 → 保区后，各区内 sc6 降序
    # 不做全局 sc6 重排，避免稳/保混排（BUG-02旧逻辑已移除）
    all_safe_bao = ([(r, '稳') for r in safe_sel_sorted] +
                    [(r, '保') for r in bao_sel_sorted])
    # 规则3：冲区排序 —— 质量优先，同质量内 sc6 高→低
    # 平行志愿按①→⑩投档，好年景多所院校同时达标时录取①号志愿；
    # 故应把层次最高（school_lv 最小）的学校排①，好年景优先进最好层次的院校。
    # 同层次内再按 sc6 高→低（更难录的排前），使志愿梯度合理。
    # 旧逻辑按 gmin_diff 升序（最近最易排①），好年景反而进质量低的学校。
    rush_sel_sorted = sorted(rush_sel[:RUSH_N],
                             key=lambda r: (r.get('school_lv', 6), -(r.get('sc6') or 0)))

    # 每个池子携带正确的 tp，不依赖位置推断
    tagged = [(r, '冲') for r in rush_sel_sorted] + all_safe_bao
    plan_vols = [{**r, 'tp': tp, 'vol_idx': i+1} for i,(r,tp) in enumerate(tagged)]

    # 【吉林平行志愿录取核心规则】
    # 1. 服从调剂=True（强制）→ 永不退档。退档仅发生在不服从调剂时。
    # 2. 录取逻辑：考生分 ≥ 提档线 → 提档 → 按①→⑥顺序匹配专业 → 首个达线即录取。
    # 3. 若考生分不达①-⑥任何一个 → 调剂触发 → 服从=True则录入组内其他任意专业（含未填报的）。
    # 4. ⑤⑥的作用：保底锚点。确保考生分 ≥ sc6，则至少录入⑥（主动选择），不触发调剂到未知专业。
    # 5. 因此 sc6 ≤ 考生分 是选组条件，非防退档，而是防调剂到未填报专业。
    for v in plan_vols:
        top6_v = v.get('top6', [])
        v['diaoji'] = True          # 服从调剂，强制 True，保证永不退档
        n_maj = len(top6_v)

        # 排除专业告警：组内含用户明确排除的专业，无论调剂是否触发均需告知
        excl = v.get('excl_in_group', [])
        if excl:
            v['warn_excl_major'] = True
            v['warn_msg_excl'] = f"组内含排除专业：{'、'.join(excl[:4])}{'等' if len(excl)>4 else ''}"
        else:
            v['warn_excl_major'] = False
            v['warn_msg_excl'] = ''

        # BUG-05：⑤⑥位冷门锚点检查（调剂到冷门比调剂到未填报更可控，但仍需提示）
        cold_anchors = [m['name'] for m in top6_v[4:6] if m.get('kind') == 'cold']
        if cold_anchors:
            v['warn_cold_anchor'] = True
            v['warn_msg_cold'] = (
                f"⑤或⑥保底专业含就业差方向：{'、'.join(cold_anchors)}，"
                f"触发调剂时将录入该专业，请确认是否可接受"
            )

    return {
        'plan_vols': plan_vols, 'all_results': all_results,
        'stats': {
            'total_cands': len(rows), 'rush_cands': len(rush_cands),
            'safe_cands': len(safe_cands), 'bao_cands': len(bao_cands),
            'plan_count': len(plan_vols), 'student_rank': student_rank,
        },
        'profile': profile,
        # 候选池暴露给优化器使用
        '_rush_cands': rush_cands,
        '_safe_cands': safe_cands,
        '_bao_cands':  bao_cands,
        '_rush_sort_key': rush_sort_key,
        '_sort_key':      sort_key,
    }


def build_tiqian(profile: dict) -> dict:
    """
    提前批志愿查询（吉林省）
    ─────────────────────────────────────────────────────────────────
    提前批与本科批完全独立，不占本科批 40 个名额。
    吉林省提前批包含：A段（公费师范、军队、公安等）、B段（部分特殊高校）。
    本函数返回考生分数区间内可报考的提前批专业组，供参考，不自动写入志愿表。
    ─────────────────────────────────────────────────────────────────
    """
    score      = int(profile['score'])
    ke_lei     = profile.get('ke_lei', '物理')
    batch      = profile.get('batch', 'A')        # 'A' / 'B' / 'all'
    score_lo   = profile.get('score_lo', score - 80)
    score_hi   = profile.get('score_hi', score + 20)
    student_province = profile.get('student_province', '吉林')

    df = load_raw_df()

    # 批次过滤
    if batch == 'A':
        batches = ['提前批A段']
    elif batch == 'B':
        batches = ['提前批B段']
    else:
        batches = ['提前批A段', '提前批B段']

    _syd = df['生源地'] if '生源地' in df.columns else None
    sub = df[
        (df['年份'] == 2025) &
        (df['科类'] == ke_lei) &
        (df['批次'].isin(batches)) &
        (_syd == student_province if _syd is not None else True)
    ].copy()

    sub['s25'] = pd.to_numeric(sub['最低分'], errors='coerce')
    sub['s24'] = pd.to_numeric(sub['最低分_1'], errors='coerce')

    # 分数范围过滤（以院校专业组最低分为基准）
    grp_min = sub.groupby('院校专业组代码')['s25'].min().rename('gmin')
    sub = sub.join(grp_min, on='院校专业组代码')
    sub = sub[(sub['gmin'] >= score_lo) & (sub['gmin'] <= score_hi)]

    # 按专业组聚合
    result = []
    seen = set()
    for _, row in sub.sort_values('gmin', ascending=False).iterrows():
        gc = row['院校专业组代码']
        if gc in seen:
            continue
        seen.add(gc)
        # 获取该组全部专业
        majors_sub = sub[sub['院校专业组代码'] == gc]
        majors = sorted(
            [{'name': str(r['专业名称']),
              's25': int(r['s25']) if pd.notna(r['s25']) else None,
              's24': int(r['s24']) if pd.notna(r['s24']) else None}
             for _, r in majors_sub.iterrows() if pd.notna(r['s25'])],
            key=lambda m: -(m['s25'] or 0)
        )
        gmin = row['gmin']
        diff = round(float(gmin) - score) if pd.notna(gmin) else None
        result.append({
            'gcode':    gc,
            'school':   str(row['院校名称']),
            'batch':    str(row['批次']),
            'city':     str(row['城市']),
            'province': str(row['所在省']),
            'gmin25':   int(float(gmin)) if pd.notna(gmin) else None,
            'diff':     diff,
            'lv':       str(row.get('院校标签', '')),
            'majors':   majors,
            'note': (
                '⚠️ 提前批：录取或退档后均不影响本科批志愿执行，可放心填报。'
                '注意体检/政审/视力等附加要求。'
            ),
        })

    return {
        'tiqian_vols': result,
        'count': len(result),
        'batch': batch,
        'ke_lei': ke_lei,
        'score': score,
        'score_range': [score_lo, score_hi],
    }


# ── 各省可用批次配置 ────────────────────────────────────────────────
# key: 数据库中 batch 列的精确值; label: 前端标签显示; default: 默认选中
# slots: 该批次志愿个数上限（0=不限/由引擎决定）
_PROVINCE_BATCHES = {
    '吉林': [
        {'key': '提前批A段',  'label': '提前批A段', 'slots': 10, 'type': 'tiqian'},
        {'key': '提前批B段',  'label': '提前批B段', 'slots': 10, 'type': 'tiqian'},
        {'key': '本科批',     'label': '本科批',    'slots': 40, 'type': 'benke', 'default': True},
        {'key': '专科批',     'label': '专科批',    'slots': 40, 'type': 'zhuanke'},
    ],
    '河北': [
        {'key': '本科提前批A段', 'label': '提前批A段', 'slots': 10, 'type': 'tiqian'},
        {'key': '本科提前批B段', 'label': '提前批B段', 'slots': 10, 'type': 'tiqian'},
        {'key': '本科提前批C段', 'label': '提前批C段', 'slots': 10, 'type': 'tiqian'},
        {'key': '本科批',        'label': '本科批',    'slots': 96, 'type': 'benke', 'default': True},
        {'key': '专科批',        'label': '专科批',    'slots': 96, 'type': 'zhuanke'},
        {'key': '专科提前批',    'label': '专科提前批', 'slots': 10, 'type': 'zhuanke'},
    ],
    '辽宁': [
        {'key': '本科提前批', 'label': '提前批',  'slots': 10, 'type': 'tiqian'},
        {'key': '本科批',     'label': '本科批',  'slots': 112, 'type': 'benke', 'default': True},
        {'key': '专科批',     'label': '专科批',  'slots': 112, 'type': 'zhuanke'},
        {'key': '专科提前批', 'label': '专科提前批', 'slots': 10, 'type': 'zhuanke'},
    ],
    '重庆': [
        {'key': '本科提前批A段', 'label': '提前批A段', 'slots': 10, 'type': 'tiqian'},
        {'key': '本科提前批B段', 'label': '提前批B段', 'slots': 10, 'type': 'tiqian'},
        {'key': '本科批',        'label': '本科批',    'slots': 96, 'type': 'benke', 'default': True},
        {'key': '专科批',        'label': '专科批',    'slots': 96, 'type': 'zhuanke'},
        {'key': '专科提前批',    'label': '专科提前批', 'slots': 10, 'type': 'zhuanke'},
    ],
    '山东': [
        {'key': '提前批',  'label': '提前批',  'slots': 10, 'type': 'tiqian'},
        {'key': '一段线',  'label': '一段(本科)', 'slots': 96, 'type': 'benke', 'default': True},
        {'key': '二段线',  'label': '二段(专科)', 'slots': 96, 'type': 'zhuanke'},
    ],
    '浙江': [
        {'key': '一段线',  'label': '一段(本科)', 'slots': 80, 'type': 'benke', 'default': True},
        {'key': '二段线',  'label': '二段(专科)', 'slots': 80, 'type': 'zhuanke'},
    ],
}

def get_province_batches(province: str) -> list:
    """返回指定省份的可用批次列表"""
    return _PROVINCE_BATCHES.get(province, [
        {'key': '本科批', 'label': '本科批', 'slots': 40, 'type': 'benke', 'default': True},
    ])


# ── 各省志愿上限配置（专业+学校直填模式）──────────────────────────────
# 格式: {省份: (total, rush, stable, safe, mode)}
# mode: 'direct'=专业+学校; 'group96'=院校+专业组但96个（河北）
_DIRECT_PROV_CFG = {
    '河北': {'total': 96,  'rush': 20, 'stable': 52, 'safe': 24, 'mode': 'group96',
             'ke_split': True,  'multi_round': False, 'invest_ratio': 1.05},
    '辽宁': {'total': 112, 'rush': 28, 'stable': 56, 'safe': 28, 'mode': 'direct',
             'ke_split': True,  'multi_round': False, 'invest_ratio': 1.05},
    '重庆': {'total': 96,  'rush': 24, 'stable': 48, 'safe': 24, 'mode': 'direct',
             'ke_split': True,  'multi_round': False, 'invest_ratio': 1.05},
    '山东': {'total': 96,  'rush': 24, 'stable': 48, 'safe': 24, 'mode': 'direct',
             'ke_split': False, 'multi_round': True,  'invest_ratio': 1.05,
             'rounds': 3},
    '浙江': {'total': 80,  'rush': 20, 'stable': 40, 'safe': 20, 'mode': 'direct',
             'ke_split': False, 'multi_round': True,  'invest_ratio': 1.0,
             'rounds': 2},
}


def build_plan_direct(profile: dict) -> dict:
    """
    专业+学校直填模式规划引擎（辽宁/重庆/山东/浙江/河北）

    与 build_plan() 的区别：
    - 不使用 major_group，直接从 major_direct 表加载数据
    - 每个志愿单位 = 1所学校 × 1个专业（无专业组中间层）
    - 山东/浙江：ke_split=False，不按物理/历史分科，统一排名
    - 浙江：invest_ratio=1.0（精确1:1投档，无调剂）
    - 河北：虽使用专业组模式，但数据在 major_direct，按冲稳保选96条
    """
    from engine.db import load_direct_df

    score    = int(profile['score'])
    province = profile.get('student_province', '吉林')
    cfg      = _DIRECT_PROV_CFG.get(province)
    if cfg is None:
        raise ValueError(f"build_plan_direct 不支持省份: {province!r}")

    ke_lei   = profile.get('ke_lei', '物理')
    slope    = float(profile.get('slope', 150.0))
    student_rank = max(1, int(7806 + (585 - score) * slope))

    # 加载该省数据
    df = load_direct_df(province)
    if df.empty:
        return {'plan_vols': [], 'stats': {}, 'profile': profile,
                'mode': 'direct', 'province_cfg': cfg}

    # ── 过滤年份 ──
    df2025 = df[df['年份'] == 2025].copy()

    # ── 过滤批次 ──
    batch = profile.get('batch', None)
    if batch and '批次' in df2025.columns:
        df2025 = df2025[df2025['批次'] == batch]
    elif '批次' in df2025.columns:
        # 默认：只保留本科批相关（排除专科批、提前批）
        _BATCH_OK = {'本科批', '一段线'}
        df2025 = df2025[df2025['批次'].isin(_BATCH_OK)]

    # ── 科类过滤（山东/浙江不分科，其他按物理/历史过滤）──
    if cfg['ke_split'] and '科类' in df2025.columns:
        df2025 = df2025[df2025['科类'] == ke_lei]

    # ── 只保留公办（与主引擎一致）──
    if '公私性质' in df2025.columns:
        df2025 = df2025[df2025['公私性质'] == '公办']

    # ── 分数有效行 ──
    df2025['s25'] = pd.to_numeric(df2025['最低分'], errors='coerce')
    df2025 = df2025[df2025['s25'].notna()].copy()

    if df2025.empty:
        return {'plan_vols': [], 'stats': {}, 'profile': profile,
                'mode': 'direct', 'province_cfg': cfg}

    # ── 用户偏好筛选 ──
    target_kw  = profile.get('target_kw', [])
    exclude_kw = profile.get('exclude_kw', [])
    excl_ne    = profile.get('exclude_northeast', False)
    prov_incl  = profile.get('province_include', [])
    prov_excl  = profile.get('province_exclude', [])
    fee_max    = profile.get('fee_max')

    # 排除关键词
    if exclude_kw:
        for kw in exclude_kw:
            mask = df2025['专业名称'].str.contains(kw, na=False)
            df2025 = df2025[~mask]

    # 排除东北（如果生源地本身就是东北省份，此选项一般不启用）
    _NE_PROVS = {'黑龙江', '吉林', '辽宁'}
    if excl_ne and '城市' in df2025.columns:
        # 通过 school 表的 province 字段或城市判断
        pass  # 直填模式数据已是本省，跳过

    # 学费上限
    if fee_max and '学费' in df2025.columns:
        _fee = pd.to_numeric(df2025['学费'], errors='coerce')
        df2025 = df2025[_fee.isna() | (_fee <= fee_max)]

    # ── 目标专业加分（优先排序）──
    df2025['_tgt_score'] = 0
    if target_kw:
        for kw in target_kw:
            df2025.loc[df2025['专业名称'].str.contains(kw, na=False), '_tgt_score'] += 10

    # ── 冲稳保区间（同主引擎逻辑）──
    rush_lo   = score - 2
    rush_hi   = score + 60
    stable_lo = score - 30
    stable_hi = score - 1
    safe_lo   = score - 60
    safe_hi   = score - 31

    def _zone(row_score):
        if rush_lo <= row_score <= rush_hi:   return '冲'
        if stable_lo <= row_score <= stable_hi: return '稳'
        if safe_lo <= row_score <= safe_hi:     return '保'
        return None

    df2025['zone'] = df2025['s25'].apply(_zone)

    # ── 构建候选列表（每行 = 1个专业志愿单元）──
    def _build_rows(zone_df, zone_label, max_n):
        """从候选中选取 max_n 条，目标专业优先、同校最多选3条避免扎堆"""
        # 按目标匹配度 + 分数排序
        sorted_df = zone_df.sort_values(['_tgt_score', 's25'],
                                         ascending=[False, zone_label != '保'])
        rows = []
        school_cnt = {}
        seen = set()
        for _, r in sorted_df.iterrows():
            if len(rows) >= max_n:
                break
            sch = str(r.get('院校名称', ''))
            mj  = str(r.get('专业名称', ''))
            key = f"{sch}|{mj}"
            if key in seen:
                continue  # 去重
            seen.add(key)
            # 同校最多3条（直填模式下避免同校扎堆）
            school_cnt[sch] = school_cnt.get(sch, 0) + 1
            if school_cnt[sch] > 3:
                continue
            def _s(val, default=''):
                """Clean NaN/None to empty string"""
                if val is None or (isinstance(val, float) and pd.isna(val)):
                    return default
                return str(val)
            rows.append({
                'school':     sch,
                'major':      mj,
                'batch':      _s(r.get('批次')),
                'city':       _s(r.get('城市')),
                'ke_lei':     _s(r.get('科类')),
                'subj_req':   _s(r.get('选科要求'), '不限'),
                'plan_count': r.get('计划人数'),
                'tuition':    r.get('学费'),
                's25':        float(r['s25']),
                'diff':       round(float(r['s25']) - score, 1),
                'tags':       _s(r.get('院校标签')),
                'city_level': _s(r.get('城市等级')),
                'school_lv':  _s(r.get('院校层级')),
                'ruanke_rank': r.get('软科排名'),
                'tp':         zone_label,
            })
        return rows

    rush_df   = df2025[df2025['zone'] == '冲']
    stable_df = df2025[df2025['zone'] == '稳']
    safe_df   = df2025[df2025['zone'] == '保']

    # ── 根据批次 slots 动态调整冲稳保数量 ──
    _batch_info = _PROVINCE_BATCHES.get(province, [])
    _batch_slots = next((b['slots'] for b in _batch_info if b['key'] == batch), cfg['total'])
    if _batch_slots >= cfg['total']:
        RUSH_N   = cfg['rush']
        STABLE_N = cfg['stable']
        SAFE_N   = cfg['safe']
    else:
        # 等比缩放：冲25% + 稳50% + 保25%
        RUSH_N   = max(1, round(_batch_slots * 0.25))
        STABLE_N = max(1, round(_batch_slots * 0.50))
        SAFE_N   = max(1, _batch_slots - RUSH_N - STABLE_N)

    rush_rows   = _build_rows(rush_df,   '冲', RUSH_N)
    stable_rows = _build_rows(stable_df, '稳', STABLE_N)
    safe_rows   = _build_rows(safe_df,   '保', SAFE_N)

    plan_vols = []
    for i, r in enumerate(rush_rows + stable_rows + safe_rows, 1):
        plan_vols.append({**r, 'vol_idx': i})

    # 多轮说明（山东/浙江）
    round_note = ''
    if cfg.get('multi_round'):
        rounds = cfg.get('rounds', 2)
        round_note = (f'本方案为第1次填报（共{rounds}次）；'
                      f'未录取可在第2次重新填报，适当降低目标。')

    return {
        'plan_vols':    plan_vols,
        'mode':         'direct',
        'province_cfg': cfg,
        'stats': {
            'total_cands':  len(df2025),
            'rush_cands':   len(rush_df),
            'stable_cands': len(stable_df),
            'safe_cands':   len(safe_df),
            'plan_count':   len(plan_vols),
            'student_rank': student_rank,
            'rush_n': RUSH_N, 'stable_n': STABLE_N, 'safe_n': SAFE_N,
        },
        'profile':    profile,
        'round_note': round_note,
    }


def mc_simulate(plan_vols, N=10000, seed=42, bias_lo=0, bias_hi=0, noise_pct=3.5,
                student_rank=7806, student_score=None):
    """
    平行志愿蒙特卡洛仿真 —— 双层噪声模型（v4，基于10072个专业组历史数据标定）

    平行志愿录取规则：
    1. 志愿组按顺序（①→㊵）尝试；第一个录取后，后续志愿全部作废。
    2. 组内专业按①→⑥顺序匹配：考生分 >= 专业当年分数线 → 录取到该专业，停止。
    3. 若考生分仅达cold保底专业（⑤⑥）→ 服从调剂，录入该专业（diaoji=True）。
    4. 若考生分低于组内所有专业 → 该组不录取，继续下一组。

    双层噪声模型参数标定（吉林省2023-2025历史数据）：
    - 真实年度变动倍数：均值=0.9922，标准差=0.0337，P5=0.9328，P95=1.0419
    - 真实年度分数绝对偏差：均值11.25分，中位7分，P90=27分
    - 年度因子 f：均值偏移0.992（历史平均每年微降0.8%），noise_pct=3.5% → U(0.957, 1.027)
    - 院校因子 g ~ U(1-σ, 1+σ)，σ=2%：院校独立波动（招生计划、专业调整等）
    - 合并 f_eff = f×g ∈ [0.938, 1.048]，|偏差|均值≈11.5分 ≈ 真实11.25分 ✓
    - 降级模式：student_score=None 时回落旧 rank-based 模型（兼容性保留）。
    """
    # 冲区志愿数（与 build_plan.RUSH_N 保持一致，用于 rush_rate 统计）
    RUSH_N      = 10
    # 院校级独立波动幅度：±2%（招生计划变动、专业调整等校内因子）
    SCHOOL_SIGMA = 0.02
    # 年度均值偏移：历史数据显示平均每年微降0.8%（考生扩招/难度波动）
    YEAR_BIAS   = -0.8   # 等效 bias_lo += 0.8（分数线平均微降）

    # 年度因子范围，叠加历史偏移（f<1=降线=乐观，f>1=涨线=悲观）
    lo_f = 1.0 + (bias_lo + YEAR_BIAS) / 100 - noise_pct / 100
    hi_f = 1.0 + (bias_hi + YEAR_BIAS) / 100 + noise_pct / 100

    # 旧 rank-based 参数（降级备用）
    lo_r = 1 - bias_hi / 100 - noise_pct / 100
    hi_r = 1 - bias_lo / 100 + noise_pct / 100

    rng          = random.Random(seed)
    n            = len(plan_vols)
    admit_counts = [0] * n
    major_hits   = [dict() for _ in range(n)]

    use_score = (student_score is not None)

    for _ in range(N):
        # ── 全局年度因子（系统性涨跌，所有院校同向）
        f = rng.uniform(lo_f, hi_f) if use_score else rng.uniform(lo_r, hi_r)

        for vi, v in enumerate(plan_vols):
            top6     = v.get('top6', v.get('intent', []))[:6]
            admitted = False

            if use_score:
                # ── 分数比较模型（双层噪声）────────────────────────
                # 全部有历史分数的专业（含cold保底，用于提档线）
                all_scored = [m for m in top6
                              if m.get('s25') and pd.notna(m['s25'])]  # 排除 NaN/0
                if not all_scored:
                    continue

                # 院校级独立因子（每所院校每轮独立抽样，打破100%相关）
                g     = rng.uniform(1.0 - SCHOOL_SIGMA, 1.0 + SCHOOL_SIGMA)
                f_eff = f * g   # 该院校该年有效分数线倍数

                # 提档线 = 组内最低专业分（含cold）× 综合因子
                sc6_thresh = all_scored[-1]['s25'] * f_eff
                if student_score < sc6_thresh:
                    continue   # 未达提档线，该组不录

                # 非cold专业优先匹配（①→⑥降序）
                valid   = [m for m in all_scored if m.get('kind') != 'cold']
                matched = False
                for m in valid:
                    if student_score >= m['s25'] * f_eff:
                        name = m.get('name', '')
                        admit_counts[vi] += 1
                        major_hits[vi][name] = major_hits[vi].get(name, 0) + 1
                        admitted = matched = True
                        break

                # 通过提档线但未匹配任何非cold专业 → 服从调剂到cold保底专业
                if not matched:
                    name = all_scored[-1].get('name', 'cold')
                    admit_counts[vi] += 1
                    major_hits[vi][name] = major_hits[vi].get(name, 0) + 1
                    admitted = True

            else:
                # ── 降级：rank-based 模型（兼容旧接口）───────────
                gr = v.get('gmin_rank') or 999999
                if gr * f < student_rank:
                    continue
                for m in top6:
                    if m.get('kind') == 'cold':
                        continue
                    mr = m.get('rank25')
                    if not mr:
                        continue
                    if mr * f >= student_rank:
                        admit_counts[vi] += 1
                        name = m.get('name', '')
                        major_hits[vi][name] = major_hits[vi].get(name, 0) + 1
                        admitted = True
                        break

            if admitted:
                break   # 平行志愿规则：已录取，后续志愿全部作废

    rates = [c / N for c in admit_counts]
    LVQ   = {1: 100, 2: 80, 3: 65, 4: 55, 5: 45, 6: 30}
    CRQ   = {1: 20,  2: 15, 3: 10, 4: 5}
    exp_q = sum(rates[i] * (LVQ.get(v['school_lv'], 30) + CRQ.get(v['city_rank'], 5))
                for i, v in enumerate(plan_vols))
    top_majors = []
    for mhd in major_hits:
        if mhd:
            top = max(mhd, key=mhd.get)
            top_majors.append({'name': top, 'count': mhd[top], 'rate': round(mhd[top] / N, 4)})
        else:
            top_majors.append(None)
    return {
        'rates':        [round(r, 4) for r in rates],
        'admit_counts': admit_counts,
        'major_hits':   major_hits,
        'top_majors':   top_majors,
        'exp_q':        round(exp_q, 2),
        'total_rate':   round(sum(rates), 4),
        'rush_rate':    round(sum(rates[:RUSH_N]), 4),   # 冲区前 RUSH_N 个志愿
        'N':            N,
    }



def optimize_plan(build_result: dict, max_rounds: int = 10,
                  mc_n: int = 8000, noise_pct: float = 3.5,
                  seed: int = 42,
                  locked_codes: set = None,
                  excluded_schools: set = None) -> dict:
    """
    多轮MC迭代优化 —— 对话框"不重复模式"的完整实现 v3。

    三类操作（按顺序每轮执行）：
    ① 冲区 - 替换：gmin_rank 不可达的死志愿 → 换可达且更贴近分数的
    ② 稳/保区 - 升级：在 sc6 ±5 分窗口内，若有质量更高（lv 更好）的院校 → 换
    ③ 稳/保区 - 去堆叠：若多个志愿 sc6 差 ≤ 3 分（扎堆），把排名靠后的换到
                         更低的 sc6 档位，拉开梯度覆盖范围

    "不重复"：每次替换记录到 blacklist，同一位置不再尝试同一学校。
    """
    import copy

    plan_vols     = [dict(v) for v in build_result['plan_vols']]
    rush_pool     = list(build_result.get('_rush_cands', []))
    safe_pool     = list(build_result.get('_safe_cands', []))
    bao_pool      = list(build_result.get('_bao_cands',  []))
    rush_sort_key = build_result.get('_rush_sort_key')
    sort_key      = build_result.get('_sort_key')
    profile       = build_result.get('profile', {})

    # 历史恢复后候选池为空 → 从 profile 重建
    if not rush_pool and not safe_pool and not bao_pool and profile:
        rebuilt = build_plan(profile)
        rush_pool     = list(rebuilt.get('_rush_cands', []))
        safe_pool     = list(rebuilt.get('_safe_cands', []))
        bao_pool      = list(rebuilt.get('_bao_cands',  []))
        rush_sort_key = rush_sort_key or rebuilt.get('_rush_sort_key')
        sort_key      = sort_key or rebuilt.get('_sort_key')
        # 用重建结果补齐 stats
        if 'stats' not in build_result or not build_result['stats']:
            build_result['stats'] = rebuilt['stats']

    student_rank  = build_result.get('stats', {}).get('student_rank', 7806)
    score         = int(profile.get('score', 585))
    noise         = noise_pct / 100.0

    LVQ = {1: 100, 2: 80, 3: 65, 4: 55, 5: 45, 6: 30}  # 层次质量分

    used_codes   = {v['gcode']  for v in plan_vols}
    used_schools = {v['school'] for v in plan_vols}
    # 用户约束：锁定的专业组代码 & 排除的院校名称
    _locked   = set(locked_codes)   if locked_codes   else set()
    _excluded = set(excluded_schools) if excluded_schools else set()
    # 从候选池中移除被排除院校
    rush_pool = [r for r in rush_pool if r['school'] not in _excluded]
    safe_pool = [r for r in safe_pool if r['school'] not in _excluded]
    bao_pool  = [r for r in bao_pool  if r['school'] not in _excluded]
    # 每个位置的黑名单（已试过、不再重复）
    blacklist = [set() for _ in plan_vols]
    for i, v in enumerate(plan_vols):
        blacklist[i].add(v['gcode'])
    # 被排除院校如果已在当前方案里 → 强制加入blacklist，第1轮必然被替换
    for i, v in enumerate(plan_vols):
        if v['school'] in _excluded and v['gcode'] not in _locked:
            blacklist[i].add(v['gcode'])   # 让优化器第1轮就替换它

    history = []

    def run_mc(vols):
        return mc_simulate(vols, N=mc_n, seed=seed,
                           noise_pct=noise_pct, student_rank=student_rank,
                           student_score=score)

    # ── 预处理：强制替换被排除院校（在进入优化循环前无条件执行）──────
    if _excluded:
        used_codes   = {v['gcode']  for v in plan_vols}
        used_schools = {v['school'] for v in plan_vols}
        for v in plan_vols:
            if v['school'] in _excluded and v['gcode'] not in _locked:
                used_codes.discard(v['gcode'])
                used_schools.discard(v['school'])
        pre_vols = [dict(v) for v in plan_vols]
        for i, v in enumerate(pre_vols):
            if v['school'] not in _excluded: continue
            if v['gcode'] in _locked: continue
            tp   = v['tp']
            pool = rush_pool if tp=='冲' else (safe_pool if tp=='稳' else bao_pool)
            s_key= rush_sort_key if tp=='冲' else sort_key
            for cand in sorted(pool, key=s_key):
                if cand['gcode']  in used_codes:    continue
                if cand['school'] in used_schools:  continue
                if cand['school'] in _excluded:     continue
                pre_vols[i] = {**cand, 'tp': tp, 'vol_idx': v['vol_idx']}
                used_codes.discard(v['gcode']);    used_codes.add(cand['gcode'])
                used_schools.discard(v['school']); used_schools.add(cand['school'])
                blacklist[i].add(cand['gcode'])
                break
        plan_vols = pre_vols   # 用预处理后的方案作为优化起点

    mc = run_mc(plan_vols)
    history.append({
        'round': 0, 'label': '初始方案',
        'plan_vols': copy.deepcopy(plan_vols),
        'mc': mc, 'changes': [],
    })

    prev_eq = mc['exp_q']
    stall   = 0

    for rnd in range(1, max_rounds + 1):
        new_vols = [dict(v) for v in plan_vols]
        changes  = []
        # 每轮开始重置 used_codes/used_schools 为当前最优方案的状态（防止跨轮污染）
        used_codes   = {v['gcode']  for v in plan_vols}
        used_schools = {v['school'] for v in plan_vols}

        # ── ① 冲区：替换不可达死志愿 ─────────────────────────────
        for i, v in enumerate(new_vols):
            if v['tp'] != '冲': continue
            if v['gcode'] in _locked: continue              # 🔒 用户锁定，跳过
            gr = v.get('gmin_rank') or 0
            if gr * (1 + noise) >= student_rank: continue   # 本身可达，跳过

            for cand in sorted(rush_pool, key=rush_sort_key):
                if cand['gcode']  in used_codes:    continue
                if cand['school'] in used_schools:  continue
                if cand['gcode']  in blacklist[i]:  continue
                if cand.get('gmin_rank', 0) * (1 + noise) < student_rank: continue
                old_school = v['school']
                new_vols[i] = {**cand, 'tp': '冲', 'vol_idx': v['vol_idx']}
                used_codes.discard(v['gcode']);    used_codes.add(cand['gcode'])
                used_schools.discard(v['school']); used_schools.add(cand['school'])
                blacklist[i].add(cand['gcode'])
                changes.append({'vol_idx': i+1, 'tp': '冲',
                                'old_school': old_school, 'old_gmin': v['gmin25'],
                                'new_school': cand['school'], 'new_gmin': cand['gmin25']})
                break

        # ── ② 稳/保区：质量升级（sc6 ±5 窗口内换更高层次院校）──────
        for i, v in enumerate(new_vols):
            if v['tp'] not in ('稳', '保'): continue
            if v['gcode'] in _locked: continue              # 🔒 用户锁定，跳过
            cur_sc6 = v.get('sc6') or 0
            cur_lv  = v.get('school_lv', 6)
            pool    = safe_pool if v['tp'] == '稳' else bao_pool

            for cand in sorted(pool, key=sort_key):  # 已按质量从高到低排
                if cand['gcode']  in used_codes:    continue
                if cand['school'] in used_schools:  continue
                if cand['gcode']  in blacklist[i]:  continue
                cand_sc6 = cand.get('sc6') or 0
                if cand_sc6 > score: continue
                if v['tp'] == '保' and score - cand_sc6 < 10: continue
                # 质量必须更高，且 sc6 在 ±5 分窗口内（避免梯度破坏）
                if cand.get('school_lv', 6) >= cur_lv: continue
                if abs(cand_sc6 - cur_sc6) > 5: continue

                old_school = v['school']
                new_vols[i] = {**cand, 'tp': v['tp'], 'vol_idx': v['vol_idx']}
                used_codes.discard(v['gcode']);    used_codes.add(cand['gcode'])
                used_schools.discard(v['school']); used_schools.add(cand['school'])
                blacklist[i].add(cand['gcode'])
                changes.append({'vol_idx': i+1, 'tp': v['tp'],
                                'old_school': old_school, 'old_gmin': v['gmin25'],
                                'new_school': cand['school'], 'new_gmin': cand['gmin25']})
                break

        # ── ③ 稳/保区：去堆叠（sc6 差 ≤ 3 的相邻志愿，推低后者扩大覆盖）
        for tp_group in ('稳', '保'):
            idxs = [i for i,v in enumerate(new_vols) if v['tp'] == tp_group]
            if len(idxs) < 2: continue
            pool = safe_pool if tp_group == '稳' else bao_pool
            for k in range(len(idxs) - 1):
                ia, ib = idxs[k], idxs[k+1]
                va, vb = new_vols[ia], new_vols[ib]
                if va['gcode'] in _locked or vb['gcode'] in _locked: continue  # 🔒 锁定
                sc6a = va.get('sc6') or 0
                sc6b = vb.get('sc6') or 0
                if abs(sc6a - sc6b) > 3: continue   # 已有足够间隔

                # 把 ib 换成一个 sc6 更低（约 -8~-15 分）的院校
                target_sc6 = sc6b - 10
                if target_sc6 < score - 50: break

                for cand in sorted(pool, key=sort_key):
                    if cand['gcode']  in used_codes:    continue
                    if cand['school'] in used_schools:  continue
                    if cand['gcode']  in blacklist[ib]: continue
                    cand_sc6 = cand.get('sc6') or 0
                    if cand_sc6 > score: continue
                    if tp_group == '保' and score - cand_sc6 < 10: continue
                    if abs(cand_sc6 - target_sc6) > 5: continue   # 要落在目标区间

                    old_school = vb['school']
                    new_vols[ib] = {**cand, 'tp': tp_group, 'vol_idx': vb['vol_idx']}
                    used_codes.discard(vb['gcode']);    used_codes.add(cand['gcode'])
                    used_schools.discard(vb['school']); used_schools.add(cand['school'])
                    blacklist[ib].add(cand['gcode'])
                    changes.append({'vol_idx': ib+1, 'tp': tp_group,
                                    'old_school': old_school, 'old_gmin': vb['gmin25'],
                                    'new_school': cand['school'], 'new_gmin': cand['gmin25']})
                    break

        plan_vols = new_vols
        mc        = run_mc(plan_vols)
        history.append({
            'round': rnd,
            'label': f'第{rnd}轮优化' if changes else f'第{rnd}轮（已收敛）',
            'plan_vols': copy.deepcopy(plan_vols),
            'mc': mc, 'changes': changes,
        })

        eq_now = mc['exp_q']
        stall  = stall + 1 if abs(eq_now - prev_eq) < 0.5 else 0
        prev_eq = eq_now
        if stall >= 2 or not changes:
            break

    best = max(history, key=lambda h: h['mc']['exp_q'])
    return {
        'plan_vols':  best['plan_vols'],
        'mc':         best['mc'],
        'history':    history,
        'best_round': best['round'],
        'stats':      build_result['stats'],
        'profile':    profile,
    }

def export_excel(plan_result: dict, mc_result: dict, out_path: str):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.workbook.properties import CalcProperties

    profile=plan_result['profile']; plan_vols=plan_result['plan_vols']
    rates=mc_result.get('rates',[0]*len(plan_vols))
    top_mjs=mc_result.get('top_majors',[None]*len(plan_vols))
    major_hits=mc_result.get('major_hits',[{}]*len(plan_vols))
    N_mc=mc_result.get('N',1)

    wb=Workbook()
    # Office 2019 兼容：关闭 full-calc-on-load，设置兼容模式
    wb.properties.lastModifiedBy = ''
    thin=Side(style='thin',color='00BDBDBD'); bdr=Border(left=thin,right=thin,top=thin,bottom=thin)
    # PatternFill 颜色必须用完整 8 位 ARGB（前两位为 FF = 不透明）
    FILL={
        '冲': PatternFill('solid', fgColor='FFFFF3E0'),
        '稳': PatternFill('solid', fgColor='FFE3F2FD'),
        '保': PatternFill('solid', fgColor='FFE8F5E9'),
        'hdr':PatternFill('solid', fgColor='FF1565C0'),
    }

    ws1=wb.active; ws1.title='志愿总表'
    score=profile.get('score','?'); ke=profile.get('ke_lei','物理')
    tgt='|'.join(profile.get('target_kw',[])[:6])

    ws1.merge_cells('A1:R1'); ws1['A1']=f'吉林省高考志愿规划表  {score}分·{ke}选科·{tgt}'
    ws1['A1'].font=Font(bold=True,size=13,color='FF1A237E')
    ws1['A1'].fill=PatternFill('solid',fgColor='FFE8EAF6')
    ws1['A1'].alignment=Alignment(horizontal='center',vertical='center'); ws1.row_dimensions[1].height=26

    ws1.merge_cells('A2:R2')
    ws1['A2']=(f"排除：{'|'.join(profile.get('exclude_kw',[])[:4]) or '无'}  ·  "
               f"排除东北：{'是' if profile.get('exclude_northeast') else '否'}  ·  "
               f"MC总录取率：{mc_result.get('total_rate',0):.1%}  ·  "
               f"冲命中率：{mc_result.get('rush_rate',0):.1%}  ·  "
               f"期望质量：{mc_result.get('exp_q',0):.1f}")
    ws1['A2'].font=Font(size=9,color='FF555555')
    ws1['A2'].fill=PatternFill('solid',fgColor='FFF5F5F5')
    ws1['A2'].alignment=Alignment(horizontal='left',vertical='center'); ws1.row_dimensions[2].height=18

    hdrs=['序号','梯度','院校名称','城市','省份','层次','专业组代码','gmin25','sc6','安全',
          '专业①','专业②','专业③','专业④','专业⑤','专业⑥','MC命中率','最高命中专业']
    widths=[5,5,22,8,8,7,14,7,7,7,18,18,18,18,18,18,9,18]
    for ci,(h,w) in enumerate(zip(hdrs,widths),1):
        c=ws1.cell(3,ci,h); c.font=Font(color='FFFFFFFF',bold=True,size=9)
        c.fill=FILL['hdr']; c.alignment=Alignment(horizontal='center',vertical='center'); c.border=bdr
        ws1.column_dimensions[get_column_letter(ci)].width=w
    ws1.row_dimensions[3].height=18; ws1.freeze_panes='A4'

    _MUNICIPALITIES = {'北京', '上海', '天津', '重庆'}
    for i,v in enumerate(plan_vols):
        row=i+4; tp=v['tp']; rate=rates[i] if i<len(rates) else 0; top=top_mjs[i] if i<len(top_mjs) else None
        sc6v=v.get('sc6'); safe_str='✅安全' if v.get('safe') else ('⚠️注意' if sc6v else '❓')
        top6=v.get('top6',[])[:6]
        def _safe_s(v): return int(v) if isinstance(v, (int,float)) and pd.notna(v) else '?'
        specs=[f"{m['name']}({_safe_s(m['s25'])})" for m in top6]
        while len(specs)<6: specs.append('')
        # 直辖市城市名修正：city 字段存储的是区名（海淀区/闵行区），用 province 替代
        _raw_city = v.get('city', '')
        _province = v.get('province', '')
        _city_display = (_province if (_raw_city and ('区' in _raw_city or '县' in _raw_city)
                         and _province in _MUNICIPALITIES) else _raw_city)
        vals=[i+1,tp,v['school'],_city_display,_province,v.get('lv_label','?'),v['gcode'],
              int(v['gmin25']) if v.get('gmin25') else '',int(sc6v) if sc6v else '',safe_str]+specs+[f"{rate:.1%}",top['name'] if top else '']
        for ci,val in enumerate(vals,1):
            c=ws1.cell(row,ci,val); c.fill=FILL[tp]; c.border=bdr
            c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
            c.font=Font(size=9,bold=(ci==3))
        ws1.row_dimensions[row].height=36

    ws2=wb.create_sheet('专业分数明细')
    hdrs2=['志愿序','梯度','院校','城市','层次','专业名称','类型','2025分','2024分','差值','MC命中率']
    widths2=[5,5,20,8,7,28,7,8,8,10,10]
    for ci,(h,w) in enumerate(zip(hdrs2,widths2),1):
        c=ws2.cell(1,ci,h); c.font=Font(color='FFFFFFFF',bold=True,size=9)
        c.fill=FILL['hdr']; c.border=bdr; c.alignment=Alignment(horizontal='center')
        ws2.column_dimensions[get_column_letter(ci)].width=w
    ws2.freeze_panes='A2'; row2=2; score_int=int(score)
    for i,v in enumerate(plan_vols):
        mhd=major_hits[i] if i<len(major_hits) else {}
        for m in v.get('majors',[]):
            s25=m.get('s25'); s24=m.get('s24'); kind=m.get('kind','other')
            _s25_ok = isinstance(s25, (int, float)) and pd.notna(s25)
            _s24_ok = isinstance(s24, (int, float)) and pd.notna(s24)
            diff=round(s25-score_int,0) if _s25_ok else None
            bg='C8E6C9' if kind=='target' else 'FFE0B2' if kind=='cold' else 'FFFFFF'
            cnt=mhd.get(m['name'],0); hr=f"{cnt/N_mc:.1%}" if cnt>0 else ''
            vals2=[i+1,v['tp'],v['school'],v['city'],v.get('lv_label','?'),m['name'],kind,
                   int(s25) if _s25_ok else '',int(s24) if _s24_ok else '',
                   (f"+{int(diff)}" if diff and diff>0 else int(diff)) if diff is not None else '',hr]
            for ci,val in enumerate(vals2,1):
                c=ws2.cell(row2,ci,val); c.fill=PatternFill('solid',fgColor=bg); c.border=bdr
                c.alignment=Alignment(horizontal='center',vertical='center')
                fc='C62828' if (diff and diff>0) else '1B5E20' if (diff and diff<0) else '000000'
                c.font=Font(size=9,color=fc)
            ws2.row_dimensions[row2].height=16; row2+=1

    ws3=wb.create_sheet('退档规则说明')
    ws3.column_dimensions['A'].width=22; ws3.column_dimensions['B'].width=60
    rules=[('【吉林省高考志愿填报核心规则】',''),('',''),
           ('志愿结构','本科批最多40个专业组，每组最多6个专业'),('本方案','40志愿（冲10+稳20+保10）'),('',''),
           ('【退档机制（核心！）】',''),
           ('档案提取条件','组最低分 ≤ 考生分 → 提档'),('录取条件','考生分 ≥ 某专业2025分 → 录取'),
           ('退档触发','提档 + 所有填报专业分>考生分 + 不服从调剂 = 退档 → 后续全废！'),
           ('历史冲击型（安全）','组最低>考生分 → 不提档 → 后续正常执行，失败零损失'),('',''),
           ('【三梯度说明】',''),
           ('⚡冲志愿','历史冲击型；gmin>考生分；意向专业有望降线命中'),
           ('✅稳志愿','sc6≤考生分；第6专业分≤考生分，服从调剂保底'),
           ('🛡保志愿','考生分远超sc6；第1专业即可进，零风险'),('',''),
           ('【填报铁律】',''),
           ('铁律1','全部勾选「服从专业调剂」'),
           ('铁律2','专业从高到低排列：最想要的填①，最低保底填⑥'),
           ('铁律3','⑥位填分数≤考生分的可接受专业作调剂锚')]
    for ri,(a,b) in enumerate(rules,1):
        ws3.cell(ri,1,a).font=Font(bold=bool(a and not b),size=10,color='FF1A237E' if a.startswith('【') else '000000')
        ws3.cell(ri,2,b).font=Font(size=9)
        if a.startswith('【'): ws3.cell(ri,1).fill=PatternFill('solid',fgColor='FFE8EAF6')
        ws3.row_dimensions[ri].height=18

    # ── 规则4：提前批 A 段独立 Sheet（Instructions核心规则）──────
    # 提前批与本科批完全独立，不占40个本科批名额
    ws4 = wb.create_sheet('提前批A段说明')
    ws4.column_dimensions['A'].width = 24
    ws4.column_dimensions['B'].width = 62
    # 标题行
    ws4.merge_cells('A1:B1')
    ws4['A1'] = '提前批A段志愿填报说明（公费师范生等）'
    ws4['A1'].font = Font(bold=True, size=13, color='FF1A237E')
    ws4['A1'].fill = PatternFill('solid', fgColor='FFE3F2FD')
    ws4['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws4.row_dimensions[1].height = 26
    adv_rules = [
        ('【关键规则】', ''),
        ('独立批次', '提前批A段与本科批完全独立，不占本科批40个专业组名额'),
        ('录取不影响本科批', '提前批录取或退档，均不影响本科批志愿的执行'),
        ('单独规划', '本方案为本科批规划；如需报考公费师范/军队院校/公安院校，须单独制定提前批方案'),
        ('', ''),
        ('【常见提前批A段类型】', ''),
        ('公费师范生', '教育部直属6所师范大学（北师大/华东师大/陕师大/西南大学/华中师大/东北师大）'),
        ('军队院校', '解放军各院校，招飞/国防生等特殊项目'),
        ('公安院校', '中国人民公安大学、中国人民警察大学等'),
        ('航海类院校', '大连海事大学等部分涉密或特殊类专业'),
        ('', ''),
        ('【填报建议】', ''),
        ('可以填', '若有意向，大胆填报，录取了比本科批更稳；未录取本科批照常投档'),
        ('退档无损', '提前批退档后，系统自动进入本科批流程，两者互不干扰'),
        ('分开备案', '本规划表仅覆盖本科批（40个志愿）；提前批另外准备，单独存档'),
    ]
    for ri, (a, b) in enumerate(adv_rules, 2):
        ca = ws4.cell(ri, 1, a)
        cb = ws4.cell(ri, 2, b)
        ca.font = Font(bold=a.startswith('【'), size=10,
                       color='FF1A237E' if a.startswith('【') else '000000')
        cb.font = Font(size=9)
        if a.startswith('【'):
            ca.fill = PatternFill('solid', fgColor='FFE3F2FD')
            cb.fill = PatternFill('solid', fgColor='FFE3F2FD')
        ws4.row_dimensions[ri].height = 18

    # 重排 sheet 顺序：志愿总表 > 专业分数明细 > 提前批A段说明 > 退档规则说明
    wb.move_sheet('提前批A段说明', offset=-(len(wb.worksheets) - 3))

    wb.save(out_path)
    return out_path


def export_excel_direct(plan_result: dict, out_path: str):
    """直填模式（辽宁/重庆/山东/浙江/河北）Excel导出"""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    profile = plan_result['profile']
    plan_vols = plan_result['plan_vols']
    prov_cfg = plan_result.get('province_cfg', {})
    prov = profile.get('student_province', '?')

    wb = Workbook()
    wb.properties.lastModifiedBy = ''
    thin = Side(style='thin', color='00BDBDBD')
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    FILL = {
        '冲': PatternFill('solid', fgColor='FFFFF3E0'),
        '稳': PatternFill('solid', fgColor='FFE3F2FD'),
        '保': PatternFill('solid', fgColor='FFE8F5E9'),
        'hdr': PatternFill('solid', fgColor='FF1565C0'),
    }

    ws = wb.active
    ws.title = '志愿总表'
    score = profile.get('score', '?')
    ke = profile.get('ke_lei', '综合')
    tgt = '|'.join(profile.get('target_kw', [])[:6])
    total = prov_cfg.get('total', len(plan_vols))

    ws.merge_cells('A1:L1')
    ws['A1'] = f'{prov}高考志愿规划表（专业+学校直填）  {score}分·{ke}·{tgt}'
    ws['A1'].font = Font(bold=True, size=13, color='FF1A237E')
    ws['A1'].fill = PatternFill('solid', fgColor='FFE8EAF6')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:L2')
    round_note = plan_result.get('round_note', '')
    ws['A2'] = (f"模式：专业+学校直填  ·  志愿上限：{total}个  ·  "
                f"排除：{'|'.join(profile.get('exclude_kw',[])[:4]) or '无'}  ·  "
                f"{round_note}")
    ws['A2'].font = Font(size=9, color='FF555555')
    ws['A2'].fill = PatternFill('solid', fgColor='FFF5F5F5')
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 18

    hdrs = ['序号', '梯度', '院校名称', '专业名称', '城市', '层次',
            '25年最低分', '差值', '科类', '选科要求', '学费', '招生计划']
    widths = [5, 5, 22, 28, 10, 8, 10, 8, 8, 12, 8, 8]
    for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
        c = ws.cell(3, ci, h)
        c.font = Font(color='FFFFFFFF', bold=True, size=9)
        c.fill = FILL['hdr']
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = bdr
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 18
    ws.freeze_panes = 'A4'

    for i, v in enumerate(plan_vols):
        row = i + 4
        tp = v['tp']
        s25 = v.get('s25')
        diff = v.get('diff')
        _s25_display = int(s25) if isinstance(s25, (int, float)) and pd.notna(s25) else '?'
        _diff_display = ''
        if isinstance(diff, (int, float)) and pd.notna(diff):
            _diff_display = f"+{int(diff)}" if diff > 0 else str(int(diff))
        tuition = v.get('tuition')
        _fee = int(tuition) if isinstance(tuition, (int, float)) and pd.notna(tuition) else ''
        plan_count = v.get('plan_count')
        _pc = int(plan_count) if isinstance(plan_count, (int, float)) and pd.notna(plan_count) else ''

        vals = [i + 1, tp, v.get('school', ''), v.get('major', ''),
                v.get('city', ''), v.get('school_lv', ''),
                _s25_display, _diff_display,
                v.get('ke_lei', ''), v.get('subj_req', ''),
                _fee, _pc]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row, ci, val)
            c.fill = FILL.get(tp, FILL['稳'])
            c.border = bdr
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.font = Font(size=9, bold=(ci in (3, 4)))
        ws.row_dimensions[row].height = 22

    # 规则说明 sheet
    ws2 = wb.create_sheet('填报规则说明')
    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 60
    mode_name = prov_cfg.get('mode', 'direct')
    ke_split = prov_cfg.get('ke_split', False)
    multi_round = prov_cfg.get('multi_round', False)
    rules = [
        (f'【{prov}高考志愿填报规则】', ''),
        ('', ''),
        ('志愿模式', f'专业+学校直填（每条志愿 = 1所学校 + 1个专业）'),
        ('志愿上限', f'本科批最多 {total} 个志愿'),
        ('科类分类', f'{"物理类/历史类分开投档" if ke_split else "不分科，全省统一排序（综合）"}'),
        ('多轮填报', f'{"是 — 落榜后可参加后续批次填报" if multi_round else "否 — 单次填报"}'),
        ('投档比例', f'{prov_cfg.get("invest_ratio", 1.05):.0%}'),
        ('', ''),
        ('【直填模式核心要点】', ''),
        ('无组内调剂', '每条志愿只对应1个专业，录取即为该专业，不存在"服从调剂"'),
        ('退档风险低', '不会出现"提档后调剂到不想去的专业"的情况'),
        ('顺序投档', '系统按志愿顺序逐一检索，一旦投档成功则停止'),
        ('', ''),
        ('【冲稳保策略】', ''),
        ('⚡冲志愿', '历年最低分高于考生分0-15分的院校专业，有冲击可能'),
        ('✅稳志愿', '历年最低分在考生分-20到+0范围内，录取把握大'),
        ('🛡保志愿', '历年最低分低于考生分20分以上，确保有学上'),
    ]
    for ri, (a, b) in enumerate(rules, 1):
        ca = ws2.cell(ri, 1, a)
        cb = ws2.cell(ri, 2, b)
        ca.font = Font(bold=a.startswith('【'), size=10,
                       color='FF1A237E' if a.startswith('【') else '000000')
        cb.font = Font(size=9)
        if a.startswith('【'):
            ca.fill = PatternFill('solid', fgColor='FFE8EAF6')
            cb.fill = PatternFill('solid', fgColor='FFE8EAF6')
        ws2.row_dimensions[ri].height = 18

    wb.save(out_path)
    return out_path
