"""
实验班数据库模块
从 data/shiyanban.xlsx 加载实验班信息，构建：
  SYBAN_MAP[(院校名称, 实验班名称)] = frozenset[分流专业名]
用于 planner.py 的 classify() 函数识别"实验班包含目标专业"的情形。
"""
import os, re, pickle, sys
from collections import defaultdict

# PyInstaller 打包路径兼容（与 planner.py 保持一致）
if getattr(sys, 'frozen', False):
    _BASE = os.path.dirname(sys.executable)
else:
    _BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_XLSX_PATH   = os.path.join(_BASE, 'data', 'shiyanban.xlsx')
_CACHE_PATH  = os.path.join(_BASE, 'data', 'syban_cache.pkl')
_CACHE_VER   = 2          # 改版本号可强制重建缓存

_syban_cache: dict | None = None   # {(school, cls_name): frozenset[major]}


def _extract_majors(full_name: str) -> list[str]:
    """从实验班全称的括号内容中提取分流专业名称列表。
    例：'工科试验班(信息)(计算机科学与技术、软件工程、自动化)'
    → ['计算机科学与技术', '软件工程', '自动化']
    """
    if not full_name:
        return []
    # 提取所有括号内内容
    parts = re.findall(r'[（(]([^）)]+)[）)]', str(full_name))
    majors = []
    for p in parts:
        # 含顿号/逗号才视为专业列表
        if '、' in p or '，' in p:
            for item in re.split(r'[、，,]', p):
                item = item.strip()
                # 过滤：长度合理、不含括号内嵌说明、不含明显非专业关键词
                if (2 <= len(item) <= 20
                        and '(' not in item and '（' not in item
                        and not any(x in item for x in ('详情', '收费', '培养', '学期', '担任', '任选'))):
                    majors.append(item)
    return majors


def load_syban_map() -> dict:
    """返回 {(院校名称, 实验班名称): frozenset[分流专业]}"""
    global _syban_cache
    if _syban_cache is not None:
        return _syban_cache

    # 尝试加载 pickle 缓存
    if os.path.exists(_CACHE_PATH):
        try:
            with open(_CACHE_PATH, 'rb') as f:
                cached = pickle.load(f)
            if isinstance(cached, dict) and cached.get('_ver') == _CACHE_VER:
                _syban_cache = cached['map']
                return _syban_cache
        except Exception:
            pass

    # 从 Excel 重建
    import openpyxl
    wb = openpyxl.load_workbook(_XLSX_PATH, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    headers = [cell.value for cell in ws[3]]

    tmp: dict[tuple, set] = defaultdict(set)
    for row in ws.iter_rows(min_row=4, values_only=True):
        d = dict(zip(headers, row))
        school   = str(d.get('院校名称') or '').strip()
        cls_name = str(d.get('实验班名称') or '').strip()
        if not school or not cls_name:
            continue
        key = (school, cls_name)
        # 1. 显式分流专业字段
        fen_liu = str(d.get('分流专业') or '').strip()
        if fen_liu and fen_liu != 'None':
            tmp[key].add(fen_liu)
        # 2. 从全称括号中解析
        for m in _extract_majors(d.get('实验班全称') or ''):
            tmp[key].add(m)
    wb.close()

    _syban_cache = {k: frozenset(v) for k, v in tmp.items()}

    # 写缓存
    try:
        with open(_CACHE_PATH, 'wb') as f:
            pickle.dump({'_ver': _CACHE_VER, 'map': _syban_cache}, f)
    except Exception:
        pass

    return _syban_cache


def get_covered_majors(school: str, cls_name: str) -> frozenset:
    """返回某院校某实验班覆盖的分流专业集合（不在 map 中则返回空集）"""
    m = load_syban_map()
    return m.get((school, cls_name), frozenset())


def is_syban_target(school: str, cls_name: str, target_kw: list[str]) -> bool:
    """判断该实验班是否覆盖了至少一个目标专业关键词"""
    if not target_kw:
        return False
    covered = get_covered_majors(school, cls_name)
    return bool(covered) and any(
        any(kw in major for kw in target_kw)
        for major in covered
    )


def matching_majors(school: str, cls_name: str, target_kw: list[str]) -> list[str]:
    """返回实验班中与 target_kw 匹配的分流专业名列表（用于前端展示）"""
    covered = get_covered_majors(school, cls_name)
    return sorted(m for m in covered if any(kw in m for kw in target_kw))
