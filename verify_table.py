"""验证生成的志愿表内容"""
import sys, os; sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

_HERE    = os.path.dirname(os.path.abspath(__file__))
_default = os.path.join(_HERE, 'outputs', '大拿_605分_吉林高考志愿表.xlsx')
_path    = sys.argv[1] if len(sys.argv) > 1 else _default
wb = load_workbook(_path)
ws = wb['志愿总表']

print('序号  梯度    院校名称              城市    层次    专业①目标')
print('-'*90)
for row in ws.iter_rows(min_row=1, max_row=65, values_only=True):
    if row[0] and str(row[0]).isdigit():
        idx  = row[0]
        tp   = str(row[1] or '')
        sch  = str(row[2] or '')[:16]
        city = str(row[3] or '')
        lv   = str(row[5] or '')
        maj1 = str(row[10] or '')[:22]
        print(f'{idx:2d}  {tp:6s}  {sch:18s}  {city:6s}  {lv:6s}  {maj1}')
