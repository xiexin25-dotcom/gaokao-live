"""生成大拿605分志愿表"""
import urllib.request, json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

STUDENT = '大拿'
SCORE   = 605
RANK    = 5000
KE      = '物理+化学+生物'

# ── 1. 生成方案 ──────────────────────────────────────────────
body = json.dumps({
    'score': SCORE, 'ke_lei': '物理', 'target_kw': '',
    'select_subjects': ['化学', '生物'],
    'min_city_rank': 2, 'school_pref': 'city',
}).encode()
req = urllib.request.Request('http://localhost:5000/api/generate',
                              data=body, headers={'Content-Type': 'application/json'})
d     = json.loads(urllib.request.urlopen(req).read())
vols  = d['vols']
rates = d['mc']['rates']
mc    = d['mc']

# ── 2. 按 tp 分组排序（修复稳/保混排）───────────────────────
# 冲区：质量优先(lv升序) → 同质量内sc6降序（与planner.py保持一致）
# 稳/保区：sc6降序（全局梯度）
TP_ORDER = {'冲': 0, '稳': 1, '保': 2}
def sort_vol(x):
    v = x[0]
    tp_o = TP_ORDER.get(v['tp'], 9)
    sc6  = v.get('sc6') or 0
    lv   = v.get('school_lv', 6)
    if v['tp'] == '冲':
        return (tp_o, lv, -sc6)       # 冲区：质量(lv)优先，同层次sc6高的先
    else:
        return (tp_o, 0, -sc6)        # 稳/保区：只按sc6降序

pairs = sorted(zip(vols, rates), key=sort_vol)
vols_out  = [v for v, r in pairs]
rates_out = [r for v, r in pairs]

# ── 3. 样式辅助 ──────────────────────────────────────────────
def fill(hex6):
    return PatternFill('solid', fgColor='FF' + hex6)

F_TITLE = fill('1A237E')
F_HEAD  = fill('283593')
F_RUSH  = fill('FFF3E0')
F_SAFE  = fill('E3F2FD')
F_BAO   = fill('E8F5E9')
F_RH    = fill('E65100')
F_SH    = fill('1565C0')
F_BH    = fill('2E7D32')
F_STAT  = fill('37474F')

thin = Side(style='thin', color='FFBDBDBD')
bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

def wc(ws, row, col, val, bold=False, size=9, color='FF212121',
       bg=None, halign='center', wrap=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(bold=bold, size=size, color=color, name='微软雅黑')
    c.alignment = Alignment(horizontal=halign, vertical='center', wrap_text=wrap)
    if bg:
        c.fill = bg
    c.border = bdr
    return c

wb = Workbook()

# ════════════════════════════════════════════════════════════
#  Sheet 1 — 志愿总表
# ════════════════════════════════════════════════════════════
ws = wb.active
ws.title = '志愿总表'
ws.sheet_view.showGridLines = False

col_widths = [4, 5, 24, 8, 8, 7, 14, 8, 8, 10, 22, 22, 22, 22, 22, 22, 10, 8]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[1].height = 38
ws.row_dimensions[2].height = 20
ws.row_dimensions[3].height = 18

# 第1行 标题
ws.merge_cells('A1:R1')
t = ws['A1']
t.value = ('🎓 ' + STUDENT + '  ·  吉林省高考志愿规划表（张雪峰方法论版）'
           + '    ' + str(SCORE) + '分 · ' + KE + ' · 全省第' + str(RANK) + '名'
           + '    策略：大城市优先 · 冲稳保（10+20+10）= 40志愿')
t.font      = Font(bold=True, size=13, color='FFFFFFFF', name='微软雅黑')
t.fill      = F_TITLE
t.alignment = Alignment(horizontal='left', vertical='center')

# 第2行 MC统计
ws.merge_cells('A2:R2')
s = ws['A2']
s.value = ('【MC仿真 N=5000】  冲区命中率 ' + str(round(mc['rush_rate']*100, 1)) + '%'
           + '  ·  整体录取率 ' + str(round(mc['total_rate']*100, 1)) + '%'
           + '  ·  质量期望值 ' + str(mc['exp_q'])
           + '  ·  城市：一线33所 + 新一线7所 · 已过滤二三线城市')
s.font      = Font(size=9, color='FFFFFFFF', name='微软雅黑')
s.fill      = F_STAT
s.alignment = Alignment(horizontal='left', vertical='center')

# 第3行 表头
headers = ['序号', '梯度', '院校名称', '城市', '城市档位', '院校层次',
           '专业组代码', '最低分25', '保底分sc6', '安全度',
           '专业①目标', '专业②', '专业③', '专业④', '专业⑤', '专业⑥',
           'MC命中率', '服从调剂']
for ci, h in enumerate(headers, 1):
    wc(ws, 3, ci, h, bold=True, size=9, color='FFFFFFFF', bg=F_HEAD)

# 数据行（修复稳/保混排：先全部冲，再全部稳，再全部保）
TP_FILL = {'冲': F_RUSH, '稳': F_SAFE, '保': F_BAO}
TP_ICON = {'冲': '⚡冲', '稳': '✅稳', '保': '🛡保'}
LV_MAP  = {1: '985', 2: '211', 3: '双一流', 4: '国重点', 5: '省重点', 6: '普通'}
LABELS  = {
    '冲': '▷ 冲志愿区（①～⑩）  sc6 > 考生分，好年景冲击名校，①号优先进最高层次',
    '稳': '▷ 稳志愿区（⑪～㉚）  sc6 ≤ 分数-2，安全录取区间，服从调剂保底',
    '保': '▷ 保志愿区（㉛～㊵）  sc6 ≤ 分数-10，绝对安全保底，差距10~50分',
}
LABEL_FILL = {'冲': F_RH, '稳': F_SH, '保': F_BH}

cur_tp  = None
row_num = 4
vol_idx = 0

for v, rate in zip(vols_out, rates_out):
    tp = v['tp']

    # 插入分区标签行
    if tp != cur_tp:
        ws.row_dimensions[row_num].height = 15
        ws.merge_cells('A' + str(row_num) + ':R' + str(row_num))
        lc = ws.cell(row_num, 1, LABELS[tp])
        lc.font      = Font(bold=True, size=8, color='FFFFFFFF', name='微软雅黑')
        lc.fill      = LABEL_FILL[tp]
        lc.alignment = Alignment(horizontal='left', vertical='center')
        cur_tp   = tp
        row_num += 1

    ws.row_dimensions[row_num].height = 20
    vol_idx += 1

    sc6      = v.get('sc6')
    diff     = round(sc6 - SCORE, 0) if sc6 else None
    diff_str = ('+' + str(int(diff)) if diff and diff > 0 else str(int(diff)) if diff else '?')
    safe     = v.get('safe', False)
    safe_str = ('✅' if safe else ('⚠️' if sc6 else '❓'))
    bg       = TP_FILL.get(tp, fill('FFFFFF'))

    # 修复1：字段名是 intent6，不是 top6
    top6  = (v.get('intent6') or v.get('top6') or [])[:6]
    specs = [m['name'] + '(' + (str(int(m['s25'])) if m.get('s25') else '?') + ')'
             for m in top6]
    while len(specs) < 6:
        specs.append('')

    rate_color = 'FFE65100' if rate > 0.08 else ('FF1565C0' if rate > 0.02 else 'FF616161')

    # 修复2：直辖市(北京/上海/天津/重庆)城市字段存的是区名，改用省级名称
    MUNICIPALITIES = {'北京', '上海', '天津', '重庆'}
    raw_city = v.get('city', '')
    province = v.get('province', '')
    if raw_city and ('区' in raw_city or '县' in raw_city) and province in MUNICIPALITIES:
        city_display = province
    else:
        city_display = raw_city

    row_vals = [
        vol_idx,
        TP_ICON.get(tp, tp),
        v['school'],
        city_display,
        v.get('cr_label', '?'),
        LV_MAP.get(v['school_lv'], '?'),
        v['gcode'],
        int(v['gmin25']) if v.get('gmin25') else '',
        int(sc6) if sc6 else '',
        diff_str + ' ' + safe_str,
    ] + specs + [
        str(round(rate * 100, 1)) + '%',
        '是 ✓',
    ]

    for ci, val in enumerate(row_vals, 1):
        bold_ci = ci in (2, 3, 11)
        col_c   = 'FFFFFFFF' if ci == 18 else (rate_color if ci == 17 else 'FF212121')
        col_bg  = fill('2E7D32') if ci == 18 else bg
        wc(ws, row_num, ci, val, bold=bold_ci, size=9,
           color=col_c, bg=col_bg,
           halign='left' if ci == 3 else 'center',
           wrap=(ci >= 11))
    row_num += 1

# ════════════════════════════════════════════════════════════
#  Sheet 2 — 填报说明
# ════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('填报说明')
ws2.sheet_view.showGridLines = False
ws2.column_dimensions['A'].width = 22
ws2.column_dimensions['B'].width = 58

ws2.row_dimensions[1].height = 32
ws2.merge_cells('A1:B1')
h1 = ws2['A1']
h1.value = STUDENT + ' 志愿填报说明  ·  ' + str(SCORE) + '分  ·  吉林省新高考'
h1.font      = Font(bold=True, size=12, color='FFFFFFFF', name='微软雅黑')
h1.fill      = F_TITLE
h1.alignment = Alignment(horizontal='center', vertical='center')

notes = [
    ('【考生信息】', ''),
    ('姓名', STUDENT),
    ('总分', str(SCORE) + '分'),
    ('全省位次', '物理组第' + str(RANK) + '名'),
    ('选科', KE),
    ('填报策略', '大城市优先（一线+新一线）→ 学校层次 → 专业'),
    ('', ''),
    ('【志愿分配】', ''),
    ('冲区（①~⑩）', '10所  sc6 > 考生分  好年景冲击985/211名校'),
    ('稳区（⑪~㉚）', '20所  sc6 ∈ [分数-45, 分数-2]  安全录取区间'),
    ('保区（㉛~㊵）', '10所  sc6 ∈ [分数-50, 分数-10]  绝对安全保底'),
    ('', ''),
    ('【城市分布】', '一线城市33所 · 新一线7所 · 已过滤三线以下'),
    ('', ''),
    ('【MC仿真结果】', ''),
    ('冲区命中率', str(round(mc['rush_rate']*100,1)) + '%（好年景进985/211概率）'),
    ('整体录取率', str(round(mc['total_rate']*100,1)) + '%（至少被某所院校录取概率）'),
    ('质量期望值', str(mc['exp_q']) + '（综合层次×概率，越高越好）'),
    ('', ''),
    ('【吉林省新高考铁律】', ''),
    ('铁律1', '全部勾选【服从专业调剂】，否则退档风险极高'),
    ('铁律2', '专业组内①~⑥务必填满，避免被调剂到未填报专业'),
    ('铁律3', '平行志愿一次投档，退档后本批次所有后续志愿作废'),
    ('铁律4', '2025为吉林新高考首年，数据波动大，建议加保底志愿'),
    ('铁律5', '服从调剂=永不退档；不服从=有退档风险（本表全部勾选）'),
]

for ri, (k, v2) in enumerate(notes, 2):
    ws2.row_dimensions[ri].height = 18
    ck = ws2.cell(ri, 1, k)
    cv = ws2.cell(ri, 2, v2)
    if k.startswith('【'):
        ck.font = Font(bold=True, size=10, color='FF1A237E', name='微软雅黑')
        ck.fill = fill('E8EAF6')
        cv.fill = fill('E8EAF6')
    else:
        ck.font = Font(size=9, bold=(k != ''), color='FF424242', name='微软雅黑')
    cv.font      = Font(size=9, color='FF212121', name='微软雅黑')
    cv.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
    ck.border = bdr; cv.border = bdr

# ── 保存 ──────────────────────────────────────────────────────
out = r'C:\Users\谢欣\Downloads\大拿_605分_吉林高考志愿表_v2.xlsx'
wb.save(out)
print('OK: ' + out)
print('size: ' + str(os.path.getsize(out)) + ' bytes')
